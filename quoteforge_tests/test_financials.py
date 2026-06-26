"""Tests for financials, reconciliation export, and daily-report financials."""
from datetime import datetime
from unittest.mock import patch

from openpyxl import load_workbook

from quoteforge.etsy.financials import order_financials, summarize, month_financials


def test_revenue_and_margin_exclude_pass_through_tax():
    # REGRESSION: sale_price (Etsy grand total) INCLUDES sales tax, a pass-through
    # ($0 net). Revenue + margin must be tax-EXCLUSIVE or both are overstated by
    # the tax. The grand total is kept separately for payout reconciliation.
    etsy = {"order_id": "T1", "status": "delivered", "sale_price": 52.21,
            "tax_collected": 3.22, "shipping_collected": 5.99, "gelato_cost": 10.0}
    r = order_financials(etsy)
    assert r["sale_price"] == 52.21               # grand total kept (payout recon)
    assert r["product_revenue"] == 48.99          # 52.21 - 3.22 tax
    s = summarize([etsy])
    assert s["revenue"] == 48.99                  # tax-exclusive top line
    assert s["gross_total"] == 52.21              # grand total preserved
    # A direct order (no ACTUAL tax) must NOT be reduced by an estimate - its
    # sale_price is already the product subtotal.
    direct = {"order_id": "W1", "status": "delivered", "sale_price": 40.0,
              "gelato_cost": 10.0}
    assert order_financials(direct)["product_revenue"] == 40.0


def test_ledger_uses_one_source_of_truth(tmp_path, monkeypatch):
    # REGRESSION: build_ledger recomputed economics with ESTIMATED fees on
    # tax-inclusive revenue, disagreeing with the reconciliation summary. It now
    # uses order_financials - tax-exclusive revenue + ACTUAL fees when imported.
    import quoteforge.db.database as db
    monkeypatch.setattr(db, "DB_PATH", tmp_path / "t.db")
    monkeypatch.setattr(db, "OUTPUT_DIR", tmp_path)
    db.init_db()
    db.create_order({"order_id": "L1", "occasion": "B", "status": "delivered",
                     "sale_price": 52.21, "gelato_cost": 10.0})
    # create_order forces status='received'; move it to a BILLABLE (earned) status so
    # it counts - the ledger now excludes non-billable orders, exactly like summarize.
    db.update_order("L1", status="delivered", tax_collected=3.22, etsy_fees_actual=5.0)
    from quoteforge.etsy.ledger import build_ledger
    tot = build_ledger("all")["totals"]
    assert tot["revenue"] == 48.99                # tax excluded, matches summarize
    assert tot["etsy_fees"] == 5.0               # ACTUAL fee, not a re-estimate


# ── Per-order financials ─────────────────────────────────────────

def test_order_financials_with_real_numbers():
    fin = order_financials({"order_id": "A", "sale_price": 29.99,
                            "gelato_cost": 11.0, "status": "shipped"})
    assert fin["sale_price"] == 29.99
    assert fin["gelato_cost"] == 11.0
    assert fin["estimated"] is False
    # net profit = sale - gelato - etsy fees, all > 0
    assert 0 < fin["net_profit"] < 29.99
    assert fin["etsy_fees"] > 0
    assert fin["sales_tax_collected"] > 0


def test_order_financials_falls_back_to_defaults():
    fin = order_financials({"order_id": "B", "status": "shipped"})  # no prices
    assert fin["estimated"] is True
    from quoteforge.config import DEFAULT_SALE_PRICE, DEFAULT_GELATO_COST
    assert fin["sale_price"] == DEFAULT_SALE_PRICE
    assert fin["gelato_cost"] == DEFAULT_GELATO_COST


def test_summarize_totals():
    orders = [
        {"order_id": "1", "sale_price": 30.0, "gelato_cost": 11.0, "status": "shipped"},
        {"order_id": "2", "sale_price": 45.0, "gelato_cost": 20.0, "status": "in_production"},
    ]
    s = summarize(orders)
    assert s["order_count"] == 2
    assert s["revenue"] == 75.0
    assert s["gelato_cost"] == 31.0
    assert s["net_profit"] > 0
    assert s["net_profit"] < s["revenue"]


def test_summarize_excludes_pending():
    orders = [
        {"order_id": "1", "sale_price": 30.0, "gelato_cost": 11.0, "status": "shipped"},
        {"order_id": "2", "sale_price": 30.0, "gelato_cost": 11.0, "status": "received"},
        {"order_id": "3", "sale_price": 30.0, "gelato_cost": 11.0, "status": "error"},
    ]
    s = summarize(orders, billable_only=True)
    assert s["order_count"] == 1  # only the shipped one counts as revenue


def test_sales_tax_is_separate_from_profit():
    fin = order_financials({"order_id": "T", "sale_price": 100.0,
                            "gelato_cost": 11.0, "status": "shipped"})
    # Tax is collected/remitted by Etsy — must NOT be folded into net profit
    assert fin["sales_tax_collected"] == 7.0  # 7% of 100
    assert fin["net_profit"] < 100.0 - 11.0   # profit excludes the tax pass-through


# ── Database integration ─────────────────────────────────────────

def test_create_order_stores_prices(tmp_path):
    import quoteforge.db.database as db
    with patch.object(db, "DB_PATH", tmp_path / "t.db"), \
         patch.object(db, "OUTPUT_DIR", tmp_path):
        db.init_db()
        db.create_order({"order_id": "P-1", "recipient_name": "Emma",
                         "occasion": "Graduation", "sale_price": 34.99,
                         "gelato_cost": 12.5})
        o = db.get_order("P-1")
    assert o["sale_price"] == 34.99
    assert o["gelato_cost"] == 12.5


def test_month_financials(tmp_path):
    import quoteforge.db.database as db
    with patch.object(db, "DB_PATH", tmp_path / "t.db"), \
         patch.object(db, "OUTPUT_DIR", tmp_path):
        db.init_db()
        oid = db.create_order({"order_id": "M-1", "recipient_name": "X",
                               "occasion": "Y", "sale_price": 30.0, "gelato_cost": 11.0})
        db.update_order(oid, status="shipped")
        now = datetime.now()
        data = month_financials(now.year, now.month)
    assert data["order_count"] == 1
    assert data["revenue"] == 30.0


# ── Reconciliation Excel ─────────────────────────────────────────

def test_reconciliation_excel(tmp_path):
    import quoteforge.db.database as db
    from quoteforge.etsy.reconciliation import export_reconciliation
    with patch.object(db, "DB_PATH", tmp_path / "t.db"), \
         patch.object(db, "OUTPUT_DIR", tmp_path):
        db.init_db()
        oid = db.create_order({"order_id": "RX-1", "recipient_name": "X",
                               "occasion": "Graduation", "sale_price": 30.0,
                               "gelato_cost": 11.0})
        db.update_order(oid, status="shipped")
        now = datetime.now()
        path = export_reconciliation(now.year, now.month, tmp_path / "recon.xlsx")
    assert path.exists()
    wb = load_workbook(path)
    ws = wb["Reconciliation"]
    # Header + at least one data row + totals row
    assert ws.max_row >= 4
    # A totals row exists
    found_total = any(ws.cell(row=r, column=1).value == "TOTAL"
                      for r in range(1, ws.max_row + 1))
    assert found_total


# ── Daily report includes financials ─────────────────────────────

def test_daily_report_has_financials(tmp_path):
    import quoteforge.db.database as db
    from quoteforge.automation.emailer import build_report_html
    with patch.object(db, "DB_PATH", tmp_path / "t.db"), \
         patch.object(db, "OUTPUT_DIR", tmp_path):
        db.init_db()
        oid = db.create_order({"order_id": "D-1", "recipient_name": "X",
                               "occasion": "Y", "sale_price": 30.0, "gelato_cost": 11.0})
        db.update_order(oid, status="shipped")
        _, body = build_report_html()
    assert "Revenue" in body
    assert "Etsy Fees" in body
    assert "NET PROFIT" in body
    assert "Gelato Print Cost" in body
    assert "Sales Tax" in body
