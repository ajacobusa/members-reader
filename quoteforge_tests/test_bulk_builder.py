"""Tests for the bulk listing builder (scale the catalog)."""
from pathlib import Path

from openpyxl import load_workbook

from quoteforge.etsy.bulk_builder import build_batch, format_batch_text
from quoteforge.etsy.launch_pack import next_additions
from quoteforge import admin


def test_build_batch_creates_packages(tmp_path):
    report = build_batch(batch=10, output_dir=tmp_path)
    assert report["batch"] == 10
    assert report["seo_clean"] == 10            # all SEO valid
    # a folder + seo.txt per listing
    folders = [p for p in Path(tmp_path).iterdir() if p.is_dir()]
    assert len(folders) == 10
    assert all((f / "seo.txt").exists() for f in folders)


def test_master_excel_has_all_rows(tmp_path):
    build_batch(batch=8, output_dir=tmp_path)
    wb = load_workbook(tmp_path / "batch_seo_master.xlsx")
    ws = wb.active
    assert ws.max_row == 9                       # header + 8
    assert ws.cell(1, 3).value == "Title"


def test_all_batch_seo_is_valid(tmp_path):
    report = build_batch(batch=25, output_dir=tmp_path)
    assert all(not r["warnings"] for r in report["listings"])


def test_scaler_has_no_nonsensical_combos():
    # No "Son Mother's Day", "Daughter Father's Day", non-spouse Anniversary.
    for a in next_additions(20, 80):
        rel, occ = a["relationship"], a["occasion"]
        if occ == "Mother's Day":
            assert rel in {"Mother", "Grandmother", "Wife"}, a["title"]
        if occ == "Father's Day":
            assert rel in {"Father", "Grandfather", "Husband"}, a["title"]
        if occ == "Anniversary":
            assert rel in {"Wife", "Husband"}, a["title"]


def test_format_text(tmp_path):
    report = build_batch(batch=5, output_dir=tmp_path)
    text = format_batch_text(report)
    assert "BULK LISTING BUILDER" in text and "Master sheet" in text


# ── CLI ──────────────────────────────────────────────────────────

def test_cli_build_batch(tmp_path, capsys):
    import quoteforge.config as config
    from unittest.mock import patch
    with patch.object(config, "OUTPUT_DIR", tmp_path):
        rc = admin.main(["build-batch", "6"])
    out = capsys.readouterr().out
    assert rc == 0
    assert "BULK LISTING BUILDER" in out
    assert "6 listing package" in out


# ── Launch kit (the actual 20 launch listings) ───────────────────

def test_launch_kit_builds_all_20(tmp_path):
    from quoteforge.etsy.bulk_builder import build_launch_kit
    from quoteforge.etsy.launch_pack import LAUNCH_PACK_20
    r = build_launch_kit(with_art=False, output_dir=tmp_path)
    assert r["count"] == len(LAUNCH_PACK_20)
    assert r["seo_clean"] == r["count"]                 # all SEO valid
    # per-listing folder with seo.txt + section label
    folders = [p for p in Path(tmp_path).iterdir() if p.is_dir()]
    assert len(folders) == len(LAUNCH_PACK_20)
    sample = (folders[0] / "seo.txt").read_text(encoding="utf-8")
    assert "SECTION:" in sample and "TITLE" in sample


def test_launch_kit_writes_checklist_and_master(tmp_path):
    from quoteforge.etsy.bulk_builder import build_launch_kit
    build_launch_kit(with_art=False, output_dir=tmp_path)
    assert (tmp_path / "UPLOAD_CHECKLIST.txt").exists()
    assert (tmp_path / "batch_seo_master.xlsx").exists()
    checklist = (tmp_path / "UPLOAD_CHECKLIST.txt").read_text(encoding="utf-8")
    assert "SECTION:" in checklist


def test_cli_launch_kit_no_art(tmp_path, capsys):
    import quoteforge.config as config
    from unittest.mock import patch
    with patch.object(config, "OUTPUT_DIR", tmp_path):
        rc = admin.main(["launch-kit", "--no-art"])
    out = capsys.readouterr().out
    assert rc == 0 and "20 listing package" in out
