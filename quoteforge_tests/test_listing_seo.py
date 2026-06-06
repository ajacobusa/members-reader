"""Tests for the per-listing Etsy SEO optimizer."""
from quoteforge.etsy.listing_seo import (
    build_launch_seo, optimize_listing, validate_seo, _resolve_niche,
    _pick_tags, export_seo_excel, format_seo_text,
)
from quoteforge.etsy.launch_pack import LAUNCH_PACK_20
from quoteforge import admin


def test_all_20_listings_pass_etsy_constraints():
    bundles = build_launch_seo()
    assert len(bundles) == 20
    for b in bundles:
        assert b.warnings == [], f"#{b.listing_n}: {b.warnings}"
        assert len(b.title) <= 140
        assert len(b.tags) == 13
        assert all(len(t) <= 20 for t in b.tags)
        assert len(set(b.tags)) == 13            # all unique
        assert b.description and "HOW IT WORKS" in b.description


def test_titles_are_front_loaded_and_substantial():
    # Good Etsy titles are long (use the space) and keyword-rich.
    for b in build_launch_seo():
        assert 60 <= len(b.title) <= 140


def test_niche_resolution_is_smart():
    assert _resolve_niche("Graduation", "Graduation", "Future Nurse Gift") == "Future Nurse"
    assert _resolve_niche("Graduation", "Graduation", "Future Dentist Gift") == "Future Dentist"
    assert _resolve_niche("Mom", "Just Because", "Personalized Grandma Gift") == "Grandma Gift"
    assert _resolve_niche("Wedding", "Anniversary", "Anniversary Gift") == "Anniversary Gift"
    assert _resolve_niche("Memorial", "Memorial", "Pet Memorial") == "Pet Memorial"
    assert _resolve_niche("Daughter", "Birthday", "Daughter Birthday") == "Daughter Gifts"


def test_tags_are_unique_and_within_limit():
    tags, warns = _pick_tags("Daughter Gifts")
    assert len(tags) == 13
    assert len(set(tags)) == 13
    assert all(len(t) <= 20 for t in tags)


def test_validate_flags_bad_bundle():
    b = optimize_listing(LAUNCH_PACK_20[0])
    b.title = "x" * 150
    b.tags = b.tags[:5]
    problems = validate_seo(b)
    assert any("140" in p for p in problems)
    assert any("13" in p for p in problems)


def test_attributes_present():
    b = optimize_listing(LAUNCH_PACK_20[0])
    assert "Occasion" in b.attributes and "Recipient" in b.attributes


def test_export_excel(tmp_path):
    from openpyxl import load_workbook
    path = export_seo_excel(tmp_path / "seo.xlsx")
    assert path.exists()
    wb = load_workbook(path)
    ws = wb.active
    assert ws.max_row == 21          # header + 20 listings
    assert ws.cell(1, 1).value == "#"


def test_description_is_ascii_safe():
    # Must not contain characters that break the Windows console / plain export.
    for b in build_launch_seo():
        b.description.encode("ascii")   # raises if non-ASCII slipped in


# ── Profession coverage (all job fields, not just nursing) ───────

def test_every_profession_in_taxonomy_passes():
    from quoteforge.etsy.listing_seo import all_profession_seo
    from quoteforge.etsy.occasions import PROFESSIONS
    bundles = all_profession_seo()
    assert len(bundles) == len(PROFESSIONS)
    for b in bundles:
        assert b.warnings == [], f"{b.niche}: {b.warnings}"
        assert len(b.tags) == 13
        assert all(len(t) <= 20 for t in b.tags)
        assert len(set(b.tags)) == 13


def test_uncurated_profession_gets_specific_tags():
    from quoteforge.etsy.listing_seo import profession_seo
    b = profession_seo("Lawyer")
    assert any("lawyer" in t for t in b.tags)
    assert "lawyer" in b.title.lower()


def test_long_profession_name_uses_short_alias():
    from quoteforge.etsy.listing_seo import profession_seo
    b = profession_seo("Veterinarian")
    # 'veterinarian' is too long for most tag phrases -> 'vet' alias used.
    assert any("vet" in t for t in b.tags)
    assert all(len(t) <= 20 for t in b.tags)


def test_curated_profession_still_uses_rich_niche():
    from quoteforge.etsy.listing_seo import profession_seo
    assert profession_seo("Nurse").niche == "Future Nurse"
    assert profession_seo("Dentist").niche == "Future Dentist"


# ── Relationship + occasion blending (Dad/Mom/Son/Daughter birthday) ──

def test_birthday_listings_get_birthday_tags():
    from quoteforge.etsy.listing_seo import relationship_seo
    for rel in ["Dad", "Mom", "Son", "Daughter"]:
        b = relationship_seo(rel, "Birthday")
        joined = " ".join(b.tags)
        assert "birthday" in joined, f"{rel}: no birthday tag"
        assert "birthday" in b.title.lower()
        assert len(b.tags) == 13 and all(len(t) <= 20 for t in b.tags)


def test_each_relationship_maps_to_its_own_niche():
    from quoteforge.etsy.listing_seo import relationship_seo
    assert relationship_seo("Dad").niche == "Dad Gift"
    assert relationship_seo("Mom").niche == "Mom Gift"
    assert relationship_seo("Son").niche == "Son Gifts"
    assert relationship_seo("Daughter").niche == "Daughter Gifts"
    assert relationship_seo("Grandma").niche == "Grandma Gift"


def test_personalized_does_not_false_match_son():
    # "Personalized" contains the substring "son" — must NOT map Dad to Son Gifts.
    from quoteforge.etsy.listing_seo import _resolve_niche
    assert _resolve_niche("Dad", "Birthday", "Personalized Dad Birthday Gift") == "Dad Gift"


def test_birthday_title_not_contradictory():
    # A birthday listing must never advertise "graduation" in its title.
    from quoteforge.etsy.listing_seo import relationship_seo
    for rel in ["Dad", "Mom", "Son", "Daughter"]:
        title = relationship_seo(rel, "Birthday").title.lower()
        assert "graduation" not in title
        assert "mothers day" not in title or rel == "Mom"


# ── CLI ──────────────────────────────────────────────────────────

def test_cli_seo_summary(capsys):
    rc = admin.main(["seo"])
    out = capsys.readouterr().out
    assert rc == 0
    assert "20 launch listings" in out


def test_cli_seo_single(capsys):
    rc = admin.main(["seo", "1"])
    out = capsys.readouterr().out
    assert rc == 0
    assert "TITLE" in out and "TAGS (13)" in out


def test_description_states_frame_not_included():
    from quoteforge.etsy.listing_seo import build_launch_seo
    desc = build_launch_seo()[0].description
    assert "FRAME NOT INCLUDED" in desc
    assert "frame shown in the photos is for display only" in desc


def test_description_lists_real_product_options():
    from quoteforge.etsy.listing_seo import build_launch_seo
    from quoteforge.config import PRODUCTS
    desc = build_launch_seo()[0].description
    assert "CHOOSE YOUR PRODUCT" in desc
    # every material in the catalog is surfaced
    materials = {name.split(" ")[0] for name in PRODUCTS}
    for m in materials:
        assert m in desc
