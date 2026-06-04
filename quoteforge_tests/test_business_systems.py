"""Tests for Gelato catalog, profit calculator, webhook server, and scaling roadmap."""
import json
from pathlib import Path
from unittest.mock import patch

from quoteforge.etsy.gelato_catalog import (
    GELATO_CATALOG,
    get_by_category,
    get_best_profit_products,
    calculate_profit,
)
from quoteforge.etsy.profit_calculator import (
    calculate_order_profit,
    monthly_revenue_forecast,
    scaling_milestones,
    canva_template_roi,
    ETSY_TRANSACTION_FEE,
    ETSY_LISTING_FEE,
)
from quoteforge.automation.webhook_server import process_webhook_payload
from quoteforge.etsy.scaling_roadmap import export_scaling_roadmap
from openpyxl import load_workbook


# ── Gelato catalog ───────────────────────────────────────────────

def test_catalog_has_enough_products():
    assert len(GELATO_CATALOG) >= 20


def test_catalog_covers_all_categories():
    categories = {p.category for p in GELATO_CATALOG}
    for expected in ["poster", "canvas", "framed", "acrylic", "metal", "mug"]:
        assert expected in categories, f"Missing category: {expected}"


def test_get_by_category_filters_correctly():
    posters = get_by_category("poster")
    assert all(p.category == "poster" for p in posters)
    assert len(posters) >= 4


def test_calculate_profit_returns_correct_structure():
    # Find first poster
    poster = next(p for p in GELATO_CATALOG if p.category == "poster")
    result = calculate_profit(poster.product_id, 29.99)
    assert "net_profit" in result
    assert "margin_pct" in result
    assert "etsy_fees" in result
    assert result["profitable"] == (result["net_profit"] > 0)


def test_calculate_profit_unknown_product():
    result = calculate_profit("nonexistent_sku", 29.99)
    assert result == {}


def test_best_profit_products_all_meet_threshold():
    products = get_best_profit_products(min_profit=10.0)
    for p in products:
        assert p.profit_margin.get("mid", 0) >= 10.0


def test_all_products_have_required_fields():
    for p in GELATO_CATALOG:
        assert p.product_id
        assert p.name
        assert p.gelato_cost_usd > 0
        assert isinstance(p.suggested_price, dict)
        assert "mid" in p.suggested_price


# ── Profit calculator ────────────────────────────────────────────

def test_calculate_order_profit_basic():
    result = calculate_order_profit(29.99, 11.00)
    assert result["sale_price"] == 29.99
    assert result["gelato_cost"] == 11.00
    assert result["net_profit"] > 0
    assert 0 < result["margin_pct"] < 100


def test_calculate_order_profit_fees_add_up():
    result = calculate_order_profit(29.99, 11.00)
    expected_net = round(29.99 - 11.00 - result["total_fees"], 2)
    assert abs(result["net_profit"] - expected_net) < 0.01


def test_monthly_revenue_forecast_structure():
    result = monthly_revenue_forecast(
        avg_daily_sales=5, avg_order_value=29,
        avg_gelato_cost=11, active_listings=100
    )
    assert "monthly_orders" in result
    assert "gross_revenue" in result
    assert "net_profit" in result
    assert "roi_pct" in result
    assert result["monthly_orders"] == 150.0


def test_monthly_forecast_with_ad_spend():
    no_ads = monthly_revenue_forecast(5, 29, 11, 100)
    with_ads = monthly_revenue_forecast(5, 29, 11, 100, monthly_ad_spend=200)
    assert with_ads["net_profit"] < no_ads["net_profit"]


def test_scaling_milestones_returns_six():
    milestones = scaling_milestones()
    assert len(milestones) == 6
    listings = [m["listings"] for m in milestones]
    assert listings == sorted(listings), "Milestones not in ascending order"


def test_canva_template_roi():
    result = canva_template_roi(50)
    assert result["templates"] == 50
    assert result["total_listings"] == 1000
    assert result["revenue_per_hr_invested"] > 0


# ── Webhook server ───────────────────────────────────────────────

def test_webhook_missing_required_fields():
    result = process_webhook_payload({"customer_name": "Alice"})
    assert result["status"] == "error"
    assert "Missing" in result["message"]


def test_webhook_processes_valid_payload(tmp_path):
    payload = {
        "customer_name": "Jennifer",
        "recipient_name": "Emma",
        "relationship": "To My Daughter",
        "occasion": "Graduation",
        "scenery": "Mountains",
        "tone": "Inspirational & Motivational",
        "memory": "She worked so hard.",
        "output_style": "Custom Quote",
        "order_id": "TEST-001",
    }
    # TEST_MODE generates a mock quote; isolate the DB so the run is clean.
    with patch("quoteforge.db.database.DB_PATH", tmp_path / "t.db"), \
         patch("quoteforge.db.database.OUTPUT_DIR", tmp_path), \
         patch("quoteforge.automation.webhook_server.OUTPUT_DIR", tmp_path), \
         patch("quoteforge.automation.pipeline_orchestrator.OUTPUT_DIR", tmp_path):
        result = process_webhook_payload(payload)
    assert result["status"] == "success"
    assert result["order_id"] == "TEST-001"


def test_webhook_logs_to_json(tmp_path):
    payload = {
        "recipient_name": "Emma",
        "occasion": "Graduation",
        "order_id": "LOG-001",
    }
    captured = []

    def fake_append(entry):
        captured.append(entry)

    with patch("quoteforge.db.database.DB_PATH", tmp_path / "t.db"), \
         patch("quoteforge.db.database.OUTPUT_DIR", tmp_path), \
         patch("quoteforge.automation.webhook_server.OUTPUT_DIR", tmp_path), \
         patch("quoteforge.automation.pipeline_orchestrator.OUTPUT_DIR", tmp_path), \
         patch("quoteforge.automation.webhook_server._append_webhook_log", side_effect=fake_append):
        process_webhook_payload(payload)
    assert len(captured) >= 1
    assert captured[-1]["order_id"] == "LOG-001"


# ── Scaling roadmap ───────────────────────────────────────────────

def test_scaling_roadmap_creates_xlsx(tmp_path):
    with patch("quoteforge.etsy.scaling_roadmap.OUTPUT_DIR", tmp_path):
        path = export_scaling_roadmap(tmp_path / "roadmap.xlsx")
    assert path.exists()


def test_scaling_roadmap_has_three_sheets(tmp_path):
    with patch("quoteforge.etsy.scaling_roadmap.OUTPUT_DIR", tmp_path):
        path = export_scaling_roadmap(tmp_path / "roadmap.xlsx")
    wb = load_workbook(path)
    assert "Scaling Roadmap" in wb.sheetnames
    assert "Weekly Tracker" in wb.sheetnames
    assert "VA Task Tracker" in wb.sheetnames


def test_scaling_roadmap_has_six_phases(tmp_path):
    with patch("quoteforge.etsy.scaling_roadmap.OUTPUT_DIR", tmp_path):
        path = export_scaling_roadmap(tmp_path / "roadmap.xlsx")
    wb = load_workbook(path)
    ws = wb["Scaling Roadmap"]
    # Row 1 = title, Row 2 = headers, Rows 3-8 = 6 phases
    assert ws.max_row >= 8


def test_weekly_tracker_has_52_weeks(tmp_path):
    with patch("quoteforge.etsy.scaling_roadmap.OUTPUT_DIR", tmp_path):
        path = export_scaling_roadmap(tmp_path / "roadmap.xlsx")
    wb = load_workbook(path)
    wt = wb["Weekly Tracker"]
    # Row 1 = title, Row 2 = headers, Rows 3-54 = 52 weeks
    assert wt.max_row >= 54
