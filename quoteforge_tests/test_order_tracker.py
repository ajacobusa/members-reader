import csv
from pathlib import Path
from unittest.mock import patch

from openpyxl import load_workbook

from quoteforge.etsy.order_tracker import (
    export_order_tracker,
    QUOTE_STATUSES,
    DESIGN_STATUSES,
    GELATO_STATUSES,
    COLUMNS,
)


def _write_order_log(tmp_path: Path, rows: list[dict]) -> None:
    log = tmp_path / "order_log.csv"
    if not rows:
        return
    with log.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=rows[0].keys())
        writer.writeheader()
        writer.writerows(rows)


def test_export_creates_xlsx(tmp_path):
    with patch("quoteforge.etsy.order_tracker.OUTPUT_DIR", tmp_path):
        path = export_order_tracker(tmp_path / "tracker.xlsx")
    assert path.exists()
    assert path.suffix == ".xlsx"


def test_workbook_has_three_sheets(tmp_path):
    with patch("quoteforge.etsy.order_tracker.OUTPUT_DIR", tmp_path):
        path = export_order_tracker(tmp_path / "tracker.xlsx")
    wb = load_workbook(path)
    assert set(wb.sheetnames) == {"Orders", "Dashboard", "SEO Reference"}


def test_orders_sheet_has_correct_headers(tmp_path):
    with patch("quoteforge.etsy.order_tracker.OUTPUT_DIR", tmp_path):
        path = export_order_tracker(tmp_path / "tracker.xlsx")
    wb = load_workbook(path)
    ws = wb["Orders"]
    headers = [ws.cell(row=1, column=c).value for c in range(1, len(COLUMNS) + 1)]
    expected = [col[0] for col in COLUMNS]
    assert headers == expected


def test_orders_populated_from_log(tmp_path):
    sample_rows = [
        {
            "timestamp": "2026-06-03T10:00:00",
            "recipient_name": "Emma",
            "sender_name": "Mom",
            "relationship": "To My Daughter",
            "occasion": "Graduation",
            "scenery": "Mountains",
            "output_style": "Personal Letter",
            "saved_dir": str(tmp_path),
        }
    ]
    _write_order_log(tmp_path, sample_rows)
    with patch("quoteforge.etsy.order_tracker.OUTPUT_DIR", tmp_path):
        path = export_order_tracker(tmp_path / "tracker.xlsx")
    wb = load_workbook(path)
    ws = wb["Orders"]
    assert ws.cell(row=2, column=4).value == "Emma"    # Recipient col
    assert ws.cell(row=2, column=2).value == "Mom"     # Customer Name col


def test_dashboard_sheet_has_formulas(tmp_path):
    with patch("quoteforge.etsy.order_tracker.OUTPUT_DIR", tmp_path):
        path = export_order_tracker(tmp_path / "tracker.xlsx")
    wb = load_workbook(path)
    dash = wb["Dashboard"]
    # At least one cell should contain a COUNTIF formula
    formula_cells = [
        dash.cell(row=r, column=5).value
        for r in range(1, 40)
        if dash.cell(row=r, column=5).value
    ]
    assert any("COUNTIF" in str(v) for v in formula_cells)


def test_seo_reference_sheet_has_data(tmp_path):
    with patch("quoteforge.etsy.order_tracker.OUTPUT_DIR", tmp_path):
        path = export_order_tracker(tmp_path / "tracker.xlsx")
    wb = load_workbook(path)
    seo = wb["SEO Reference"]
    assert seo.cell(row=1, column=1).value == "Niche"
    assert seo.max_row >= 10


def test_status_constants_are_complete():
    assert "Pending" in QUOTE_STATUSES
    assert "Approved" in QUOTE_STATUSES
    assert "Not Started" in DESIGN_STATUSES
    assert "Done" in DESIGN_STATUSES
    assert "Not Uploaded" in GELATO_STATUSES
    assert "Shipped" in GELATO_STATUSES


def test_export_with_no_log_creates_blank_rows(tmp_path):
    # No order_log.csv — should still create file with blank placeholder rows
    with patch("quoteforge.etsy.order_tracker.OUTPUT_DIR", tmp_path):
        path = export_order_tracker(tmp_path / "tracker.xlsx")
    wb = load_workbook(path)
    ws = wb["Orders"]
    assert ws.max_row >= 30   # 30 blank placeholder rows
