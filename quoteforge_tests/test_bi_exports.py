"""The BI exporters build real, valid deliverables from live order data: Excel
workbooks with native charts, a Power BI star-schema CSV package + DAX + docs,
and an executive presentation PDF."""
import csv

from openpyxl import load_workbook


def _seed(tmp_path, monkeypatch):
    monkeypatch.setattr("quoteforge.db.database.DB_PATH", tmp_path / "t.db")
    monkeypatch.setattr("quoteforge.db.database.OUTPUT_DIR", tmp_path)
    from quoteforge.db import database as db
    db.init_db()
    db.create_order({"order_id": "QF-1", "etsy_order_id": "1",
                     "recipient_name": "A", "occasion": "B"})
    db.update_order("QF-1", status="delivered", vendor="gelato", channel="etsy",
                    product_type="print", sale_price=49.0, gelato_cost=12.0,
                    etsy_fees_actual=4.5, acquisition_source="etsy_search")
    return db


def test_excel_workbooks_have_tabs_and_charts(tmp_path, monkeypatch):
    _seed(tmp_path, monkeypatch)
    from quoteforge.analytics.bi_exports import build_excel_workbooks
    paths = build_excel_workbooks(tmp_path / "Excel")
    assert len(paths) == 3 and all(p.exists() for p in paths)
    fin = load_workbook(paths[0])
    assert "Fee Breakdown" in fin.sheetnames and "P&L Summary" in fin.sheetnames
    total_charts = sum(len(ws._charts) for p in paths
                       for ws in load_workbook(p).worksheets)
    assert total_charts >= 3                      # pie + line + bars


def test_powerbi_package_is_complete(tmp_path, monkeypatch):
    _seed(tmp_path, monkeypatch)
    from quoteforge.analytics.bi_exports import build_powerbi_package
    out = tmp_path / "Power BI"
    r = build_powerbi_package(out)
    names = {p.name for p in r["data"]}
    assert {"fact_orders.csv", "fact_ledger_daily.csv", "fact_fees.csv",
            "dim_date.csv", "dim_vendor.csv"} <= names
    assert r["measures"].exists() and "Net Margin %" in r["measures"].read_text()
    assert r["model"].exists() and "star schema" in r["model"].read_text().lower()
    assert r["readme"].exists()
    # fact_orders has one row per order with the expected header.
    with (out / "data" / "fact_orders.csv").open(encoding="utf-8") as fh:
        rows = list(csv.DictReader(fh))
    assert len(rows) == 1 and rows[0]["order_id"] == "QF-1"
    assert rows[0]["sale_price"] == "49.0"


def test_presentation_pdf_is_built(tmp_path, monkeypatch):
    _seed(tmp_path, monkeypatch)
    from quoteforge.analytics.bi_exports import build_presentation
    p = build_presentation(tmp_path / "deck.pdf")
    assert p.exists()
    head = p.read_bytes()[:5]
    assert head.startswith(b"%PDF")               # valid PDF header
    assert p.stat().st_size > 1500


def test_export_all_writes_everything(tmp_path, monkeypatch):
    _seed(tmp_path, monkeypatch)
    from quoteforge.analytics.bi_exports import export_all
    r = export_all(tmp_path / "Excel", tmp_path / "Power BI")
    assert len(r["excel"]) == 3
    assert r["presentation"].exists()
    assert (tmp_path / "Power BI" / "data" / "fact_orders.csv").exists()
