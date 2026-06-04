"""Tests for the seasonal campaign generator (publish-early-to-rank-first)."""
from datetime import datetime
from unittest.mock import patch

from openpyxl import load_workbook

from quoteforge.etsy.campaign import (
    recommended_publish_date, build_listing_plan, seasonal_campaign,
    export_campaign_excel, DEFAULT_LEAD_DAYS,
)
from quoteforge import admin


# ── Publish-by timing (the core "be first" logic) ───────────────

def test_publish_by_is_before_peak():
    # Planning Valentine's in early January
    now = datetime(2026, 1, 5)
    timing = recommended_publish_date("Valentine's Day", 2, now)
    assert timing["peak"] == "2026-02-14"
    # Publish-by must be ~5 weeks earlier
    assert timing["publish_by"] < timing["peak"]


def test_publish_by_uses_lead_days():
    now = datetime(2026, 5, 1)
    timing = recommended_publish_date("Christmas", 12, now)
    peak = datetime.strptime(timing["peak"], "%Y-%m-%d")
    pub = datetime.strptime(timing["publish_by"], "%Y-%m-%d")
    assert (peak - pub).days == DEFAULT_LEAD_DAYS


def test_overdue_flag_when_late():
    # It's already December but planning Christmas → past the publish window
    now = datetime(2026, 12, 10)
    timing = recommended_publish_date("Christmas", 12, now)
    assert "OVERDUE" in timing["urgency"]


def test_on_track_when_early():
    now = datetime(2026, 1, 1)
    timing = recommended_publish_date("Christmas", 12, now)
    assert timing["urgency"] == "On track"


def test_next_year_rollover():
    # Planning January while in December → peak is next year
    now = datetime(2026, 12, 15)
    timing = recommended_publish_date("New Year's Day", 1, now)
    assert timing["peak"].startswith("2027")


# ── Listing plan content ─────────────────────────────────────────

def test_listing_plan_fields():
    plan = build_listing_plan("Father's Day", 6, datetime(2026, 4, 1))
    assert plan["occasion"] == "Father's Day"
    assert len(plan["title"]) <= 140
    assert len(plan["tags"]) <= 13
    assert all(len(t) <= 20 for t in plan["tags"])  # Etsy tag limit
    assert plan["scenery"]
    assert plan["publish_by"]


def test_scenery_matches_occasion():
    assert "Sunrise" in build_listing_plan("Pet Memorial", 10)["scenery"] \
        or "Gentle" in build_listing_plan("Pet Memorial", 10)["scenery"]
    assert "Floral" in build_listing_plan("Wedding — Bride", 6)["scenery"]
    assert "Patriotic" in build_listing_plan("Veterans Day", 11)["scenery"]


# ── Campaign for a month ─────────────────────────────────────────

def test_seasonal_campaign_sorted_by_urgency():
    plans = seasonal_campaign("June", datetime(2026, 4, 1))
    assert len(plans) >= 5
    # Sorted soonest publish-by first
    days = [p["days_to_publish"] for p in plans]
    assert days == sorted(days)


def test_campaign_accepts_month_number():
    by_name = seasonal_campaign("June", datetime(2026, 4, 1))
    by_num = seasonal_campaign(6, datetime(2026, 4, 1))
    assert len(by_name) == len(by_num)


def test_campaign_excel(tmp_path):
    path = export_campaign_excel("June", tmp_path / "camp.xlsx",
                                 now=datetime(2026, 4, 1))
    assert path.exists()
    wb = load_workbook(path)
    ws = wb["Campaign"]
    headers = [ws.cell(row=2, column=c).value for c in range(1, 8)]
    assert "Publish By" in headers
    assert "Etsy Title" in headers
    assert ws.max_row > 3  # has listing rows


# ── CLI ──────────────────────────────────────────────────────────

def test_cli_campaign(tmp_path, capsys):
    with patch("quoteforge.etsy.campaign.OUTPUT_DIR", tmp_path):
        rc = admin.main(["campaign", "June"])
    out = capsys.readouterr().out
    assert rc == 0
    assert "JUNE CAMPAIGN" in out
    assert "PUBLISH BY" in out
    assert "listings planned" in out


def test_cli_campaign_bad_month(capsys):
    rc = admin.main(["campaign", "Smarch"])
    assert rc == 2
    assert "Unknown month" in capsys.readouterr().out
