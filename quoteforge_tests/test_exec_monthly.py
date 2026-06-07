"""Tests for the executive report, monthly review, and ledger Summary tab."""
import openpyxl


def _seed(tmp_path, monkeypatch):
    monkeypatch.setattr("quoteforge.db.database.DB_PATH", tmp_path / "t.db")
    monkeypatch.setattr("quoteforge.db.database.OUTPUT_DIR", tmp_path)
    monkeypatch.setattr("quoteforge.config.OUTPUT_DIR", tmp_path)
    monkeypatch.setattr("quoteforge.etsy.reconciliation.OUTPUT_DIR", tmp_path, raising=False)
    monkeypatch.setattr("quoteforge.config.TEST_MODE", True, raising=False)
    from quoteforge.db import database as db
    db.init_db()
    return db


def test_exec_report_tabs_and_summary(tmp_path, monkeypatch):
    _seed(tmp_path, monkeypatch)
    from quoteforge.etsy.exec_report import build_exec_report
    out = build_exec_report("all", out_path=tmp_path / "exec.xlsx")
    wb = openpyxl.load_workbook(out)
    assert wb.sheetnames == ["Executive Summary", "P&L (daily)", "Breakdown",
                             "Infrastructure", "Workflow"]
    assert wb["Workflow"].max_row >= 16   # header + 15 stages
    inf = wb["Infrastructure"]
    text = " ".join(str(c.value) for row in inf.iter_rows() for c in row if c.value)
    for kw in ("CURRENT STATE", "AI WORKLOAD", "FUTURE-STATE",
               "BUSINESS OPPORTUNITIES", "COST-REDUCTION"):
        assert kw in text


def test_ledger_has_summary_tab(tmp_path, monkeypatch):
    _seed(tmp_path, monkeypatch)
    from quoteforge.etsy.ledger import export_ledger_excel
    out = export_ledger_excel("all", out_path=tmp_path / "gl.xlsx")
    names = openpyxl.load_workbook(out).sheetnames
    assert names[0] == "Summary" and "General Ledger" in names


def test_monthly_review_archives_three(tmp_path, monkeypatch):
    _seed(tmp_path, monkeypatch)
    from quoteforge.automation.weekly_review import monthly_review
    r = monthly_review(email=False)
    assert len(r["archive"]) == 3
    names = [p.replace("\\", "/").split("/")[-1] for p in r["archive"]]
    assert any("reconciliation" in n for n in names)
    assert any("general_ledger" in n for n in names)
    assert any("executive_report" in n for n in names)
    # archived under cost/<YYYY>/<tag>_monthly/
    assert "_monthly" in r["archive"][0].replace("\\", "/")


def test_monthly_review_emails(tmp_path, monkeypatch):
    _seed(tmp_path, monkeypatch)
    sent = {}
    monkeypatch.setattr("quoteforge.automation.emailer._send_email",
        lambda s, b, to="", attachments=None: sent.update(att=attachments, subj=s))
    from quoteforge.automation.weekly_review import monthly_review
    monthly_review(email=True)
    assert sent.get("att") and len(sent["att"]) == 3 and "Monthly" in sent["subj"]


def test_commands_registered():
    from quoteforge import admin
    assert "exec-report" in admin.COMMANDS and "monthly-review" in admin.COMMANDS
