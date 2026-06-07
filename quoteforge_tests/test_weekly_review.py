"""Tests for the Friday business review + ledger breakdown tabs."""


def _seed(tmp_path, monkeypatch):
    monkeypatch.setattr("quoteforge.db.database.DB_PATH", tmp_path / "t.db")
    monkeypatch.setattr("quoteforge.db.database.OUTPUT_DIR", tmp_path)
    monkeypatch.setattr("quoteforge.config.TEST_MODE", True, raising=False)
    from quoteforge.db import database as db
    db.init_db()
    return db


def test_weekly_review_collects_and_summarizes(tmp_path, monkeypatch):
    _seed(tmp_path, monkeypatch)
    from quoteforge.automation.weekly_review import weekly_review, collect_metrics
    m = collect_metrics()
    assert "pnl_month" in m and "tco" in m and "counts" in m and "aov" in m
    r = weekly_review(email=False)
    assert isinstance(r["summary"], str) and r["summary"]


def test_weekly_review_emails_owner(tmp_path, monkeypatch):
    _seed(tmp_path, monkeypatch)
    sent = {}
    monkeypatch.setattr("quoteforge.automation.emailer._send_email",
                        lambda s, b, to="", attachments=None: sent.update(to=to, subj=s, att=attachments))
    from quoteforge.automation.weekly_review import weekly_review
    r = weekly_review(email=True)
    assert r.get("emailed_to") and sent["to"] == r["emailed_to"]
    assert "Friday" in sent["subj"]


def test_ledger_excel_has_breakdown_tabs(tmp_path, monkeypatch):
    db = _seed(tmp_path, monkeypatch)
    oid = db.create_order({"order_id": "X", "product_type": "canvas",
                           "channel": "etsy", "vendor": "gelato"})
    db.update_order(oid, sale_price=100.0, gelato_cost=30.0)
    from quoteforge.etsy.ledger import export_ledger_excel
    out = export_ledger_excel("all", out_path=tmp_path / "gl.xlsx")
    import openpyxl
    wb = openpyxl.load_workbook(out)
    assert {"By Channel", "By Vendor", "By Product"} <= set(wb.sheetnames)


def test_weekly_review_command_registered():
    from quoteforge import admin
    assert "weekly-review" in admin.COMMANDS


def test_cost_folder_archive_dated(tmp_path, monkeypatch):
    monkeypatch.setattr("quoteforge.db.database.DB_PATH", tmp_path / "t.db")
    monkeypatch.setattr("quoteforge.db.database.OUTPUT_DIR", tmp_path)
    monkeypatch.setattr("quoteforge.config.OUTPUT_DIR", tmp_path)
    monkeypatch.setattr("quoteforge.config.TEST_MODE", True, raising=False)
    from quoteforge.db import database as db
    db.init_db()
    from datetime import date
    from quoteforge.automation.weekly_review import weekly_review
    r = weekly_review(email=False)
    arch = r.get("archive")
    assert arch and isinstance(arch, list) and arch[0].endswith(".xlsx")
    from pathlib import Path
    p = Path(arch[0])
    assert p.exists()
    today = date.today().isoformat()
    # cost/<YYYY>/<YYYY-MM-DD>/ structure
    assert p.parent.name == today and p.parent.parent.name == today[:4]
    assert p.parent.parent.parent.name == "cost"
    assert (p.parent / f"business_review_{today}.txt").exists()


def test_email_attaches_ledger(tmp_path, monkeypatch):
    monkeypatch.setattr("quoteforge.db.database.DB_PATH", tmp_path / "t.db")
    monkeypatch.setattr("quoteforge.db.database.OUTPUT_DIR", tmp_path)
    monkeypatch.setattr("quoteforge.config.OUTPUT_DIR", tmp_path)
    monkeypatch.setattr("quoteforge.config.TEST_MODE", True, raising=False)
    from quoteforge.db import database as db
    db.init_db()
    sent = {}
    monkeypatch.setattr("quoteforge.automation.emailer._send_email",
        lambda s, b, to="", attachments=None: sent.update(att=attachments))
    from quoteforge.automation.weekly_review import weekly_review
    weekly_review(email=True)
    assert sent.get("att") and str(sent["att"][0]).endswith(".xlsx")


def test_weekly_trend_four_weeks():
    from quoteforge.etsy.ledger import weekly_trend
    t = weekly_trend(4)
    assert len(t) == 4 and all("week_of" in w and "net" in w for w in t)


def test_archive_includes_reconciliation(tmp_path, monkeypatch):
    monkeypatch.setattr("quoteforge.db.database.DB_PATH", tmp_path / "t.db")
    monkeypatch.setattr("quoteforge.db.database.OUTPUT_DIR", tmp_path)
    monkeypatch.setattr("quoteforge.config.OUTPUT_DIR", tmp_path)
    monkeypatch.setattr("quoteforge.etsy.reconciliation.OUTPUT_DIR", tmp_path, raising=False)
    monkeypatch.setattr("quoteforge.config.TEST_MODE", True, raising=False)
    from quoteforge.db import database as db
    db.init_db()
    from quoteforge.automation.weekly_review import weekly_review
    r = weekly_review(email=False)
    names = [p.split("/")[-1].split("\\")[-1] for p in r["archive"]]
    assert any("general_ledger" in n for n in names)
    assert any("reconciliation" in n for n in names)


def test_ledger_excel_has_trend_tab(tmp_path, monkeypatch):
    monkeypatch.setattr("quoteforge.db.database.DB_PATH", tmp_path / "t.db")
    monkeypatch.setattr("quoteforge.db.database.OUTPUT_DIR", tmp_path)
    from quoteforge.db import database as db
    db.init_db()
    from quoteforge.etsy.ledger import export_ledger_excel
    import openpyxl
    out = export_ledger_excel("all", out_path=tmp_path / "gl.xlsx")
    assert "Trend (4 wks)" in openpyxl.load_workbook(out).sheetnames
