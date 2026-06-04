"""Scaling roadmap Excel generator — 1,000+ listings roadmap tracker."""
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from quoteforge.config import OUTPUT_DIR

CLR = {
    "phase1": "E2EFDA",  # green — validate
    "phase2": "DEEBF7",  # blue — build
    "phase3": "FFF2CC",  # yellow — automate
    "phase4": "FCE4D6",  # orange — scale
    "phase5": "F4B8C1",  # red/pink — dominate
    "phase6": "D9D2E9",  # purple — business
    "header": "1F4E79",
    "white":  "FFFFFF",
    "alt":    "F2F2F2",
}


def _cell(ws, row, col, value="", bold=False, bg=None, fg="000000",
          center=False, wrap=False, size=10):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name="Arial", bold=bold, color=fg, size=size)
    if bg:
        c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(
        horizontal="center" if center else "left",
        vertical="center", wrap_text=wrap)
    s = Side(style="thin", color="CCCCCC")
    c.border = Border(left=s, right=s, top=s, bottom=s)
    return c


def export_scaling_roadmap(output_path: Path | None = None) -> Path:
    if output_path is None:
        output_path = OUTPUT_DIR / "QuoteForge_Scaling_Roadmap.xlsx"
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    wb = Workbook()

    # ── Sheet 1: Roadmap ─────────────────────────────────────────
    ws = wb.active
    ws.title = "Scaling Roadmap"
    ws.sheet_view.showGridLines = False

    # Title
    ws.merge_cells("A1:J1")
    t = ws["A1"]
    t.value = "QuoteForge — 1,000+ Listings Scaling Roadmap"
    t.font = Font(name="Arial", bold=True, size=16, color=CLR["white"])
    t.fill = PatternFill("solid", fgColor=CLR["header"])
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    # Column headers
    headers = ["Phase", "Timeline", "Listings", "Daily Sales",
               "Monthly Rev ($)", "Monthly Profit ($)", "Listings/Week",
               "VAs Needed", "Key Focus", "Action Items"]
    widths   = [12, 14, 10, 12, 16, 18, 14, 12, 28, 50]
    for col, (h, w) in enumerate(zip(headers, widths), 1):
        c = _cell(ws, 2, col, h, bold=True, bg=CLR["header"], fg=CLR["white"],
                  center=True, size=11)
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[2].height = 28

    phases = [
        {
            "phase": "Phase 1\nValidate",
            "timeline": "Month 1",
            "listings": 30,
            "daily_sales": 0.5,
            "monthly_rev": 450,
            "monthly_profit": 280,
            "listings_week": 8,
            "vas": 0,
            "focus": "Test 7 niches. Get first 10 sales. Learn what customers buy.",
            "actions": "• Create Etsy shop + connect Gelato\n• Build 30 listings manually\n• Focus: Daughter, Graduation, Christian, Memorial\n• No automation yet — learn the workflow",
            "color": CLR["phase1"],
        },
        {
            "phase": "Phase 2\nBuild",
            "timeline": "Month 2–3",
            "listings": 100,
            "daily_sales": 2,
            "monthly_rev": 1800,
            "monthly_profit": 1100,
            "listings_week": 15,
            "vas": 0,
            "focus": "Scale to 100 listings. Add canvas + framed. Improve SEO.",
            "actions": "• Use QuoteForge Bulk Generator for 70 new listings\n• Add canvas and framed variants to top sellers\n• Run Etsy SEO using keyword database\n• Start Pinterest account",
            "color": CLR["phase2"],
        },
        {
            "phase": "Phase 3\nAutomate",
            "timeline": "Month 3–5",
            "listings": 300,
            "daily_sales": 6,
            "monthly_rev": 5400,
            "monthly_profit": 3200,
            "listings_week": 25,
            "vas": 1,
            "focus": "Automate quote generation. Hire 1 part-time VA.",
            "actions": "• Set up Zapier → QuoteForge webhook\n• Train VA on Order Fulfillment SOP\n• Use QuoteForge bulk catalog to reach 300 listings\n• Launch Etsy ads ($50/mo budget)\n• Build 50 Canva master templates",
            "color": CLR["phase3"],
        },
        {
            "phase": "Phase 4\nScale",
            "timeline": "Month 5–8",
            "listings": 600,
            "daily_sales": 14,
            "monthly_rev": 12600,
            "monthly_profit": 7500,
            "listings_week": 35,
            "vas": 2,
            "focus": "Pinterest traffic. Etsy ads. 2 VAs. Dominate 10 niches.",
            "actions": "• 600 listings across 20+ niches\n• 2 VAs handling all order fulfillment\n• Pinterest: 20 pins/day (VA task)\n• Increase Etsy ads to $200/mo\n• Add acrylic + metal prints\n• Begin second Etsy shop (different niche)",
            "color": CLR["phase4"],
        },
        {
            "phase": "Phase 5\nDominate",
            "timeline": "Month 8–12",
            "listings": 1000,
            "daily_sales": 25,
            "monthly_rev": 22500,
            "monthly_profit": 13500,
            "listings_week": 40,
            "vas": 3,
            "focus": "1,000 listings. Open Shopify store. Wholesale inquiries.",
            "actions": "• 1,000+ listings across 3 Etsy shops\n• Launch Shopify store with Gelato integration\n• Drive traffic: Pinterest + Instagram reels + TikTok\n• Build email list (offer free print for signup)\n• Corporate/office bulk order pricing\n• 3 VAs: design, fulfillment, social media",
            "color": CLR["phase5"],
        },
        {
            "phase": "Phase 6\nBusiness",
            "timeline": "Month 12+",
            "listings": 2000,
            "daily_sales": 50,
            "monthly_rev": 45000,
            "monthly_profit": 27000,
            "listings_week": 50,
            "vas": 5,
            "focus": "Full business. Multiple shops. B2B. Passive income.",
            "actions": "• 2,000+ listings across multiple channels\n• B2B: sell wall art bundles to offices, hospitals, dental offices\n• License designs to other Gelato sellers\n• Full team: VA manager + 5 VAs\n• Revenue from: Etsy + Shopify + B2B + licensing",
            "color": CLR["phase6"],
        },
    ]

    for i, p in enumerate(phases, start=3):
        ws.row_dimensions[i].height = 80
        row_data = [
            p["phase"], p["timeline"], p["listings"], p["daily_sales"],
            p["monthly_rev"], p["monthly_profit"], p["listings_week"],
            p["vas"], p["focus"], p["actions"],
        ]
        for col, val in enumerate(row_data, 1):
            bold = col == 1
            _cell(ws, i, col, val, bold=bold, bg=p["color"],
                  center=(col <= 8), wrap=(col >= 9), size=9 if col >= 9 else 10)

    # ── Sheet 2: Weekly Action Tracker ───────────────────────────
    wt = wb.create_sheet("Weekly Tracker")
    wt.sheet_view.showGridLines = False
    wt.merge_cells("A1:F1")
    h = wt["A1"]
    h.value = "Weekly Listing Creation Tracker"
    h.font = Font(name="Arial", bold=True, size=14, color=CLR["white"])
    h.fill = PatternFill("solid", fgColor=CLR["header"])
    h.alignment = Alignment(horizontal="center", vertical="center")
    wt.row_dimensions[1].height = 30

    wk_headers = ["Week", "Target Listings", "Actual Listings", "New Revenue ($)", "Notes", "Phase"]
    wk_widths   = [8, 16, 16, 16, 35, 12]
    for col, (h_text, w) in enumerate(zip(wk_headers, wk_widths), 1):
        _cell(wt, 2, col, h_text, bold=True, bg=CLR["header"], fg=CLR["white"], center=True)
        wt.column_dimensions[get_column_letter(col)].width = w

    targets = [8, 8, 8, 6, 15, 15, 15, 15, 25, 25, 25, 25,
               35, 35, 35, 35, 40, 40, 40, 40, 40, 40, 40, 40,
               50, 50, 50, 50, 50, 50, 50, 50, 50, 50, 50, 50,
               50, 50, 50, 50, 50, 50, 50, 50, 50, 50, 50, 50, 50, 50, 50, 50]
    phases_map = ([1]*4 + [2]*8 + [3]*12 + [4]*16 + [5]*8 + [6]*4)[:52]

    for week in range(1, 53):
        alt = week % 2 == 0
        bg = CLR["alt"] if alt else CLR["white"]
        target = targets[week-1] if week <= len(targets) else 50
        phase = phases_map[week-1] if week <= len(phases_map) else 6
        _cell(wt, week + 2, 1, f"Week {week}", center=True, bg=bg)
        _cell(wt, week + 2, 2, target, center=True, bg=bg)
        _cell(wt, week + 2, 3, "", center=True, bg=bg)
        _cell(wt, week + 2, 4, f"=C{week+2}*29*0.6", center=True, bg=bg)
        _cell(wt, week + 2, 5, "", bg=bg)
        _cell(wt, week + 2, 6, f"Phase {phase}", center=True, bg=bg)

    # ── Sheet 3: VA Task Tracker ─────────────────────────────────
    va = wb.create_sheet("VA Task Tracker")
    va.sheet_view.showGridLines = False
    va.merge_cells("A1:G1")
    vh = va["A1"]
    vh.value = "Virtual Assistant Daily Task Tracker"
    vh.font = Font(name="Arial", bold=True, size=14, color=CLR["white"])
    vh.fill = PatternFill("solid", fgColor=CLR["header"])
    vh.alignment = Alignment(horizontal="center", vertical="center")
    va.row_dimensions[1].height = 30

    va_headers = ["Date", "VA Name", "Orders Fulfilled", "Listings Created",
                  "Pinterest Pins", "Hours Worked", "Notes"]
    va_widths   = [12, 16, 16, 16, 14, 12, 35]
    for col, (h_text, w) in enumerate(zip(va_headers, va_widths), 1):
        _cell(va, 2, col, h_text, bold=True, bg=CLR["header"], fg=CLR["white"], center=True)
        va.column_dimensions[get_column_letter(col)].width = w

    for row in range(3, 33):
        alt = row % 2 == 0
        bg = CLR["alt"] if alt else CLR["white"]
        for col in range(1, 8):
            _cell(va, row, col, "", bg=bg)

    wb.save(output_path)
    return output_path
