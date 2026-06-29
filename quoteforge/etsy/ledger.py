"""End-to-end general ledger - daily revenue, expenses & net profit.

Captures EVERY cost so the bottom line is true:
  + Revenue        : order sale prices
  - COGS           : Gelato print cost per order
  - Etsy fees      : transaction 6.5% + payment 3% + $0.20 listing, per order
  - API cost       : Claude (Anthropic) usage from the cost log
  - OpEx           : fixed monthly overhead (Make.com + MONTHLY_FIXED_COSTS),
                     prorated per day
  = Net profit

Reads live data so it's always current; the daily job also persists a snapshot
per day for history/trends. Money figures are USD.
"""
from __future__ import annotations

import logging
from datetime import date, datetime, timedelta

logger = logging.getLogger(__name__)


def _daily_opex() -> float:
    """Fixed monthly overhead prorated to a per-day USD figure."""
    from quoteforge.config import MONTHLY_FIXED_COSTS, USE_MAKE_COM, MAKE_COM_COST
    monthly = float(MONTHLY_FIXED_COSTS) + (MAKE_COM_COST if USE_MAKE_COM else 0.0)
    return round(monthly / 30.4, 4)


def _range_for(period: str):
    """Start/end dates for a period name (today/week/month/year/all)."""
    today = date.today()
    if period == "today":
        return today, today
    if period == "week":
        return today - timedelta(days=6), today
    if period == "month":
        return today.replace(day=1), today
    if period == "year":
        return today.replace(month=1, day=1), today
    return date(2020, 1, 1), today          # "all"


def _day_of(iso: str) -> str:
    """Date part (YYYY-MM-DD) of an ISO timestamp; today if missing."""
    try:
        return (iso or "")[:10] or date.today().isoformat()
    except Exception:  # noqa: BLE001
        return date.today().isoformat()


def build_ledger(period: str = "month") -> dict:
    """Compute the P&L by day for the period + totals."""
    from quoteforge.db.database import init_db, get_all_orders, get_api_costs
    init_db()
    start, end = _range_for(period)
    # Orders/API costs are timestamped in UTC; allow +1 day so a late-evening
    # local sale stamped "tomorrow" in UTC still counts.
    hi = (end + timedelta(days=1)).isoformat()
    days: dict[str, dict] = {}

    def _row(d):
        """Get-or-create the ledger row for a day."""
        return days.setdefault(d, {"day": d, "revenue": 0.0, "cogs": 0.0,
                                   "etsy_fees": 0.0, "api_cost": 0.0,
                                   "opex": 0.0, "net_profit": 0.0, "orders": 0})

    for o in get_all_orders(limit=100000):
        d = _day_of(o.get("created_at"))
        if not (start.isoformat() <= d <= hi):
            continue
        sale = o.get("sale_price")
        if sale in (None, 0):
            continue                          # no confirmed sale price yet
        # REFUND/RETURN PARITY: only EARNED revenue counts. A refunded / cancelled
        # order keeps a sale_price but is not billable, so it must NOT sit in the
        # daily P&L as profit (it did in the old code, while summarize() + BI already
        # excluded it - the two layers reported different net profit for the month).
        from quoteforge.etsy.financials import order_financials, BILLABLE_STATUSES
        if o.get("status") not in BILLABLE_STATUSES:
            continue
        # Use order_financials (ONE source of truth) so the ledger agrees with the
        # reconciliation summary: tax-EXCLUSIVE product revenue, and ACTUAL Etsy
        # fees when imported (was: re-estimated fees on tax-inclusive revenue, so
        # the two reporting layers disagreed once real figures were imported).
        fin = order_financials(o)
        r = _row(d)
        r["revenue"] += fin["product_revenue"]
        # COGS = Gelato product cost + Gelato's shipping charge to you.
        r["cogs"] += fin["gelato_cost"] + fin.get("gelato_shipping", 0)
        r["etsy_fees"] += fin["etsy_fees"]
        r["orders"] += 1

    for c in get_api_costs(start.isoformat(), (end + timedelta(days=1)).isoformat()):
        _row(_day_of(c.get("created_at")))["api_cost"] += float(c.get("cost_usd") or 0)

    # Misc income (affiliate commissions, wholesale, etc.) - net revenue.
    from quoteforge.db.database import get_income
    for inc in get_income(start.isoformat(), hi):
        r = _row(_day_of(inc.get("day")))
        r["revenue"] += float(inc.get("amount") or 0)

    # Prorated fixed overhead for each calendar day in range.
    opex = _daily_opex()
    cur = start
    while cur <= end:
        _row(cur.isoformat())["opex"] += opex
        cur += timedelta(days=1)

    # For 'all', don't pad thousands of empty pre-history days: start at the
    # first day that actually has revenue or orders.
    if period == "all":
        active = [d for d, r in days.items() if r["revenue"] or r["orders"]]
        first = min(active) if active else end.isoformat()
        days = {d: r for d, r in days.items() if d >= first}

    rows = []
    for d in sorted(days):
        r = days[d]
        r["net_profit"] = round(r["revenue"] - r["cogs"] - r["etsy_fees"]
                                - r["api_cost"] - r["opex"], 2)
        for k in ("revenue", "cogs", "etsy_fees", "api_cost", "opex"):
            r[k] = round(r[k], 2)
        rows.append(r)

    tot = {k: round(sum(r[k] for r in rows), 2)
           for k in ("revenue", "cogs", "etsy_fees", "api_cost", "opex", "net_profit")}
    tot["orders"] = sum(r["orders"] for r in rows)
    tot["margin_pct"] = round(tot["net_profit"] / tot["revenue"] * 100, 1) \
        if tot["revenue"] else 0.0
    return {"period": period, "start": start.isoformat(), "end": end.isoformat(),
            "days": rows, "totals": tot}


def weekly_trend(weeks: int = 4) -> list[dict]:
    """Revenue / net / orders for the last `weeks` weeks (oldest -> newest)."""
    led = build_ledger("year")
    today = date.today()
    buckets = {i: {"revenue": 0.0, "net": 0.0, "orders": 0} for i in range(weeks)}
    for r in led["days"]:
        try:
            d = date.fromisoformat(r["day"][:10])
        except Exception:  # noqa: BLE001
            continue
        delta = (today - d).days
        if 0 <= delta < weeks * 7:
            b = delta // 7
            buckets[b]["revenue"] += r["revenue"]
            buckets[b]["net"] += r["net_profit"]
            buckets[b]["orders"] += r["orders"]
    out = []
    for i in range(weeks - 1, -1, -1):
        ws = today - timedelta(days=i * 7 + 6)
        out.append({"week_of": ws.isoformat(),
                    "revenue": round(buckets[i]["revenue"], 2),
                    "net": round(buckets[i]["net"], 2),
                    "orders": buckets[i]["orders"]})
    return out


def build_breakdown(period: str = "month") -> dict:
    """Revenue / cost / profit grouped by channel, vendor, and product type."""
    from quoteforge.db.database import (
        init_db, get_all_orders, get_income)
    from quoteforge.etsy.profit_calculator import calculate_order_profit
    init_db()
    start, end = _range_for(period)
    hi = (end + timedelta(days=1)).isoformat()
    by_channel: dict[str, dict] = {}
    by_vendor: dict[str, dict] = {}
    by_product: dict[str, dict] = {}

    def _acc(bucket, key):
        """Get-or-create the accumulator row for a grouping key."""
        return bucket.setdefault(key, {"revenue": 0.0, "cost": 0.0,
                                       "net": 0.0, "orders": 0})

    for o in get_all_orders(limit=100000):
        d = _day_of(o.get("created_at"))
        if not (start.isoformat() <= d <= hi):
            continue
        sale = o.get("sale_price")
        if sale in (None, 0):
            continue
        cost = o.get("gelato_cost") or 0.0
        p = calculate_order_profit(float(sale), float(cost))
        ch = o.get("channel") or "etsy"
        vd = o.get("vendor") or "gelato"
        pt = o.get("product_type") or "print"
        for bucket, key in ((by_channel, ch), (by_vendor, vd), (by_product, pt)):
            a = _acc(bucket, key)
            a["revenue"] += p["sale_price"]
            a["cost"] += p["gelato_cost"] + p["total_fees"]
            a["net"] += p["net_profit"]
            a["orders"] += 1

    for inc in get_income(start.isoformat(), hi):
        a = _acc(by_channel, inc.get("channel") or "affiliate")
        amt = float(inc.get("amount") or 0)
        a["revenue"] += amt
        a["net"] += amt                      # commission/wholesale = net income
        a["orders"] += 1

    def _round(bucket):
        """Round all money fields in a bucket to 2 decimals (in place)."""
        for k in bucket:
            for f in ("revenue", "cost", "net"):
                bucket[k][f] = round(bucket[k][f], 2)
        return bucket
    return {"period": period,
            "by_channel": _round(by_channel),
            "by_vendor": _round(by_vendor),
            "by_product": _round(by_product)}


def format_breakdown_text(bd: dict) -> str:
    """Render the channel/vendor/product breakdown as printable console text."""
    lines = ["=" * 60, f"LEDGER BREAKDOWN ({bd['period']})", "=" * 60]
    for title, key in (("BY CHANNEL", "by_channel"),
                       ("BY VENDOR", "by_vendor"),
                       ("BY PRODUCT TYPE", "by_product")):
        lines.append(f"\n{title}:")
        rows = bd[key]
        if not rows:
            lines.append("  (no sales yet)")
            continue
        lines.append(f"  {'Key':16}{'Revenue':>10}{'Cost':>10}{'Net':>10}{'Ord':>5}")
        for k in sorted(rows, key=lambda x: -rows[x]["net"]):
            r = rows[k]
            lines.append(f"  {k[:16]:16}{r['revenue']:>10.2f}{r['cost']:>10.2f}"
                         f"{r['net']:>10.2f}{r['orders']:>5}")
    lines.append("=" * 60)
    return "\n".join(lines)


def snapshot_today() -> dict:
    """Persist today's ledger row (run daily). Returns today's figures."""
    from quoteforge.db.database import upsert_ledger_snapshot
    led = build_ledger("today")
    row = led["days"][0] if led["days"] else {
        "revenue": 0, "cogs": 0, "etsy_fees": 0, "api_cost": 0,
        "opex": _daily_opex(), "net_profit": round(-_daily_opex(), 2), "orders": 0}
    upsert_ledger_snapshot(date.today().isoformat(), row)
    return row


def format_ledger_text(led: dict) -> str:
    """Render the daily ledger plus totals as printable console text."""
    t = led["totals"]
    lines = ["=" * 70,
             f"GENERAL LEDGER  ({led['period']}: {led['start']} -> {led['end']})",
             "=" * 70,
             f"{'Date':12}{'Rev':>9}{'COGS':>9}{'Fees':>8}{'API':>7}"
             f"{'OpEx':>7}{'Net':>9}{'Ord':>5}"]
    for r in led["days"]:
        lines.append(f"{r['day']:12}{r['revenue']:>9.2f}{r['cogs']:>9.2f}"
                     f"{r['etsy_fees']:>8.2f}{r['api_cost']:>7.2f}{r['opex']:>7.2f}"
                     f"{r['net_profit']:>9.2f}{r['orders']:>5}")
    lines += ["-" * 70,
              f"{'TOTAL':12}{t['revenue']:>9.2f}{t['cogs']:>9.2f}{t['etsy_fees']:>8.2f}"
              f"{t['api_cost']:>7.2f}{t['opex']:>7.2f}{t['net_profit']:>9.2f}"
              f"{t['orders']:>5}",
              f"Net margin: {t['margin_pct']}%   (Revenue {t['revenue']:.2f} - "
              f"COGS {t['cogs']:.2f} - Fees {t['etsy_fees']:.2f} - API "
              f"{t['api_cost']:.2f} - OpEx {t['opex']:.2f})",
              "=" * 70]
    return "\n".join(lines)


def export_ledger_excel(period: str = "all", out_path=None):
    """Write the ledger to an .xlsx workbook. Returns the path."""
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from quoteforge.config import OUTPUT_DIR
    led = build_ledger(period)
    wb = Workbook()
    # Summary tab (front) - exec KPIs + cost-mix chart.
    led_t = led["totals"]
    su = wb.active
    su.title = "Summary"
    su["A1"] = f"General Ledger - Summary ({period})"
    su["A1"].font = Font(bold=True, size=14)
    kpis = [("Revenue", led_t["revenue"]), ("COGS (Gelato)", led_t["cogs"]),
            ("Etsy fees", led_t["etsy_fees"]), ("API cost", led_t["api_cost"]),
            ("Overhead", led_t["opex"]), ("Net profit", led_t["net_profit"]),
            ("Net margin %", led_t["margin_pct"]), ("Orders", led_t["orders"])]
    for i, (lbl, val) in enumerate(kpis, start=3):
        su.cell(i, 1, lbl); su.cell(i, 2, val)
    try:
        from openpyxl.chart import PieChart, Reference
        # cost rows are at A4:A7 (COGS, Etsy fees, API, Overhead)
        pie = PieChart(); pie.title = "Cost mix"
        pie.add_data(Reference(su, min_col=2, min_row=4, max_row=7))
        pie.set_categories(Reference(su, min_col=1, min_row=4, max_row=7))
        su.add_chart(pie, "D3")
    except Exception as exc:  # noqa: BLE001 - chart is cosmetic
        logger.debug("ledger cost-mix pie chart skipped: %s", exc)
    su.column_dimensions["A"].width = 18

    ws = wb.create_sheet("General Ledger")
    headers = ["Date", "Revenue", "COGS (Gelato)", "Etsy Fees", "API Cost",
               "OpEx", "Net Profit", "Orders"]
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True)
    for r in led["days"]:
        ws.append([r["day"], r["revenue"], r["cogs"], r["etsy_fees"],
                   r["api_cost"], r["opex"], r["net_profit"], r["orders"]])
    t = led["totals"]
    ws.append(["TOTAL", t["revenue"], t["cogs"], t["etsy_fees"], t["api_cost"],
               t["opex"], t["net_profit"], t["orders"]])
    for c in ws[ws.max_row]:
        c.font = Font(bold=True)

    # Breakdown tabs: by channel, vendor, product type.
    bd = build_breakdown(period)
    for title, key in (("By Channel", "by_channel"), ("By Vendor", "by_vendor"),
                       ("By Product", "by_product")):
        sh = wb.create_sheet(title)
        sh.append([title.replace("By ", ""), "Revenue", "Cost", "Net Profit", "Orders"])
        for c in sh[1]:
            c.font = Font(bold=True)
        rows = bd[key]
        for k in sorted(rows, key=lambda x: -rows[x]["net"]):
            r = rows[k]
            sh.append([k, r["revenue"], r["cost"], r["net"], r["orders"]])

    # Trend tab: last 4 weeks + a native line chart (no extra deps).
    try:
        from openpyxl.chart import LineChart, Reference
        tr = wb.create_sheet("Trend (4 wks)")
        tr.append(["Week of", "Revenue", "Net Profit", "Orders"])
        for c in tr[1]:
            c.font = Font(bold=True)
        rows = weekly_trend(4)
        for r in rows:
            tr.append([r["week_of"], r["revenue"], r["net"], r["orders"]])
        if rows:
            chart = LineChart()
            chart.title = "Revenue & Net Profit - last 4 weeks"
            chart.y_axis.title = "$"
            data = Reference(tr, min_col=2, max_col=3, min_row=1,
                             max_row=len(rows) + 1)
            cats = Reference(tr, min_col=1, min_row=2, max_row=len(rows) + 1)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            tr.add_chart(chart, "F2")
    except Exception as exc:  # noqa: BLE001 - chart is cosmetic
        logger.debug("ledger trend chart skipped: %s", exc)

    out = out_path or (OUTPUT_DIR / "general_ledger.xlsx")
    from pathlib import Path
    out = Path(out)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    return out
