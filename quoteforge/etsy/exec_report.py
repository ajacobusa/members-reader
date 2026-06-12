"""Executive report - a board-ready workbook with charts + strategy.

Tabs:
  1. Executive Summary  - headline KPIs, AI exec summary, cost-mix pie, trend line
  2. P&L (daily)        - the full ledger (all costs)
  3. Breakdown          - revenue/cost/net by channel, vendor, product (+ bar chart)
  4. Infrastructure     - current state (how it's built) + AI workload assessment
                          + future-state recommendations, business opportunities,
                          and cost-reduction opportunities

AI sections use Claude with a strong deterministic fallback (works in TEST_MODE).
"""
from __future__ import annotations

from datetime import date


def _kpis(period: str) -> dict:
    """Headline KPIs for the period: P&L totals, AOV, subscribers, reviews."""
    from quoteforge.etsy.ledger import build_ledger
    from quoteforge.db.database import (init_db, subscriber_count, review_stats,
                                        get_subscriptions)
    init_db()
    t = build_ledger(period)["totals"]
    n = t.get("orders", 0)
    return {
        "revenue": t["revenue"], "cogs": t["cogs"], "etsy_fees": t["etsy_fees"],
        "api_cost": t["api_cost"], "opex": t["opex"], "net_profit": t["net_profit"],
        "margin_pct": t["margin_pct"], "orders": n,
        "aov": round(t["revenue"] / n, 2) if n else 0.0,
        "subscribers": subscriber_count(),
        "active_subs": len(get_subscriptions("active")),
        "reviews": review_stats(),
    }


def _infra_facts() -> dict:
    """Programmatically collected current-state facts about the build."""
    facts = {}
    try:
        from quoteforge.automation.scheduler import SCHEDULED_JOBS
        facts["scheduled_jobs"] = len(SCHEDULED_JOBS)
    except Exception:  # noqa: BLE001
        facts["scheduled_jobs"] = "?"
    try:
        from quoteforge import admin
        facts["admin_commands"] = len(admin.COMMANDS)
    except Exception:  # noqa: BLE001
        facts["admin_commands"] = "?"
    try:
        from quoteforge.catalog.registry import VENDORS, list_products
        facts["vendors"] = len(VENDORS)
        facts["catalog_items"] = len(list_products())
    except Exception:  # noqa: BLE001
        pass
    import glob
    facts["test_files"] = len(glob.glob("quoteforge_tests/test_*.py"))
    return facts


CURRENT_STATE = [
    ("Language / runtime", "Python 3.x, standard-library-first; zero heavy deps"),
    ("Data store", "SQLite (WAL mode) - orders, subscribers, subscriptions, "
                   "reviews, catalog, income, api_costs, ledger snapshots"),
    ("Order intake", "Etsy poll every 10 min (+ optional /order webhook) - no "
                     "Make/Zapier dependency"),
    ("AI", "Anthropic Claude (Haiku) - quotes, gift notes, photo/vision QC, "
           "Ask Ange bot, ops & weekly/monthly reviews; all TEST_MODE-safe"),
    ("Fulfillment", "Vendor-routed: Gelato (API) + Printful/Printify (ready) + "
                    "digital/manual; live price/availability sync"),
    ("Storefront", "Etsy (cart+payment) + free GitHub Pages brand site (Ask Ange, "
                   "live preview, reviews, order-by urgency)"),
    ("Automation", "Windows Task Scheduler / cron - reports, tracking, backups, "
                   "marketing, AI reviews"),
    ("Safety", "60% margin floor, QC-before-proof, refunds always human-approved, "
               "nightly backup+push, no API keys on the static site"),
]

AI_WORKLOAD = [
    ("Quote / message generation", "per order", "~$0.01-0.02 (Haiku)"),
    ("Gift notes + subscription/welcome emails", "per event", "~$0.005"),
    ("Photo + final-product vision QC", "per order (optional)", "~$0.01"),
    ("Ask Ange (on-page)", "free - client-side KB; Claude only via hosted /ask",
     "$0 default"),
    ("Ops review / weekly / monthly reviews", "scheduled", "~$0.01 each"),
    ("Assessment", "Haiku keeps AI ~1-2% of revenue at scale; cap with "
     "batching + KB-first fallbacks (already in place).", ""),
]

FUTURE_STATE = [
    "Host the server 24/7 (Render ~$7/mo) -> live Ask Ange + webhooks.",
    "At volume (>~30 orders/day): move SQLite -> Postgres + a job queue (RQ/Celery).",
    "Turn on Printful/Printify keys to auto-route non-Gelato products.",
    "A/B test thumbnails + price once ~30+ sales exist.",
    "Add Google Analytics + Clarity (already wired) to measure funnel.",
    "Event-driven tracking via Gelato webhooks (vs 6-hourly poll) when available.",
]

BUSINESS_OPPS = [
    "Subscriptions / memberships - recurring revenue (built).",
    "B2B / wholesale (Faire) + corporate gifting landing page.",
    "Affiliate 'complete the gift' (flowers/gift cards) - passive income (built).",
    "Digital downloads - ~95% margin, instant delivery.",
    "Seasonal gift-guide SEO + Pinterest - highest-ROI traffic.",
    "Pet portraits & memorials - high emotional value, strong reviews.",
]

COST_REDUCTION = [
    "Local renderer (free) instead of paid Bannerbear - already default.",
    "Claude Haiku instead of larger models - ~1/3 the cost, fine for short copy.",
    "Native Etsy poller instead of Make.com (~$9/mo saved).",
    "Bundle/quantity discounts hold the 60% floor - protect margin, not give it away.",
    "GitHub Pages (free) for the site; only ~$15/yr domain + optional $7/mo server.",
    "Gelato volume + local production network reduces unit + shipping cost.",
]


def build_exec_report(period: str = "month", out_path=None):
    """Build the executive workbook. Returns the path."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.chart import PieChart, LineChart, BarChart, Reference
    from quoteforge.config import OUTPUT_DIR, SHOP_NAME
    from quoteforge.etsy.ledger import build_ledger, build_breakdown, weekly_trend

    k = _kpis(period)
    facts = _infra_facts()
    HDR = PatternFill("solid", fgColor="103D2E")
    BOLDW = Font(bold=True, color="FFFFFF")

    wb = Workbook()

    # ── Tab 1: Executive Summary ──
    s = wb.active
    s.title = "Executive Summary"
    s["A1"] = f"{SHOP_NAME} - Executive Summary ({period}, {date.today().isoformat()})"
    s["A1"].font = Font(bold=True, size=15, color="103D2E")
    rows = [
        ("Revenue", k["revenue"]), ("Net profit", k["net_profit"]),
        ("Net margin %", k["margin_pct"]), ("Orders", k["orders"]),
        ("Avg order value", k["aov"]), ("Subscribers", k["subscribers"]),
        ("Active subscriptions", k["active_subs"]),
        ("Reviews (count / avg)",
         f"{k['reviews'].get('count', 0)} / {k['reviews'].get('avg', 0)}"),
    ]
    s["A3"] = "KEY METRICS"; s["A3"].font = Font(bold=True)
    for i, (lbl, val) in enumerate(rows, start=4):
        s.cell(i, 1, lbl); s.cell(i, 2, val)

    # cost mix for the pie
    s["A14"] = "COST MIX"; s["A14"].font = Font(bold=True)
    cm = [("COGS (Gelato)", k["cogs"]), ("Etsy fees", k["etsy_fees"]),
          ("API (Claude)", k["api_cost"]), ("Overhead", k["opex"])]
    for i, (lbl, val) in enumerate(cm, start=15):
        s.cell(i, 1, lbl); s.cell(i, 2, max(val, 0))
    pie = PieChart(); pie.title = "Cost mix"
    pie.add_data(Reference(s, min_col=2, min_row=15, max_row=18))
    pie.set_categories(Reference(s, min_col=1, min_row=15, max_row=18))
    s.add_chart(pie, "D3")

    # AI exec summary text
    summary = _ai_exec_summary(SHOP_NAME, k, facts)
    s["A21"] = "AI EXECUTIVE SUMMARY"; s["A21"].font = Font(bold=True)
    s.merge_cells("A22:H34")
    c = s["A22"]; c.value = summary
    c.alignment = Alignment(wrap_text=True, vertical="top")
    s.column_dimensions["A"].width = 26
    s.column_dimensions["B"].width = 20

    # ── Tab 2: P&L (daily) ──
    led = build_ledger(period)
    pl = wb.create_sheet("P&L (daily)")
    pl.append(["Date", "Revenue", "COGS", "Etsy Fees", "API", "OpEx", "Net", "Orders"])
    for cc in pl[1]:
        cc.font = BOLDW; cc.fill = HDR
    for r in led["days"]:
        pl.append([r["day"], r["revenue"], r["cogs"], r["etsy_fees"],
                   r["api_cost"], r["opex"], r["net_profit"], r["orders"]])
    tr = weekly_trend(4)
    pl["J1"] = "4-week trend"; pl["J1"].font = Font(bold=True)
    pl["J2"], pl["K2"], pl["L2"] = "Week of", "Revenue", "Net"
    for i, w in enumerate(tr, start=3):
        pl.cell(i, 10, w["week_of"]); pl.cell(i, 11, w["revenue"]); pl.cell(i, 12, w["net"])
    line = LineChart(); line.title = "Revenue & Net - 4 weeks"
    line.add_data(Reference(pl, min_col=11, max_col=12, min_row=2, max_row=2 + len(tr)),
                  titles_from_data=True)
    line.set_categories(Reference(pl, min_col=10, min_row=3, max_row=2 + len(tr)))
    pl.add_chart(line, "J7")

    # ── Tab 3: Breakdown ──
    bd = build_breakdown(period)
    bk = wb.create_sheet("Breakdown")
    row = 1
    for title, key in (("By Channel", "by_channel"), ("By Vendor", "by_vendor"),
                       ("By Product", "by_product")):
        bk.cell(row, 1, title).font = Font(bold=True); row += 1
        bk.cell(row, 1, "Key"); bk.cell(row, 2, "Revenue")
        bk.cell(row, 3, "Cost"); bk.cell(row, 4, "Net"); bk.cell(row, 5, "Orders")
        for cc in bk[row]:
            cc.font = Font(bold=True)
        start = row + 1
        data = bd[key]
        for kk in sorted(data, key=lambda x: -data[x]["net"]):
            row += 1
            d = data[kk]
            bk.cell(row, 1, kk); bk.cell(row, 2, d["revenue"])
            bk.cell(row, 3, d["cost"]); bk.cell(row, 4, d["net"])
            bk.cell(row, 5, d["orders"])
        if key == "by_channel" and row >= start:
            bar = BarChart(); bar.title = "Net profit by channel"
            bar.add_data(Reference(bk, min_col=4, min_row=start - 1, max_row=row),
                         titles_from_data=True)
            bar.set_categories(Reference(bk, min_col=1, min_row=start, max_row=row))
            bk.add_chart(bar, "G2")
        row += 2

    # ── Tab 4: Infrastructure & Roadmap ──
    inf = wb.create_sheet("Infrastructure")

    def _section(ws, r, title, pairs_or_list):
        """Write a titled section of pairs or bullets; return the next free row."""
        ws.cell(r, 1, title).font = Font(bold=True, size=12, color="103D2E"); r += 1
        if pairs_or_list and isinstance(pairs_or_list[0], tuple):
            for item in pairs_or_list:
                ws.cell(r, 1, item[0]).font = Font(bold=True)
                ws.cell(r, 2, " | ".join(str(x) for x in item[1:]))
                r += 1
        else:
            for item in pairs_or_list:
                ws.cell(r, 2, "- " + item); r += 1
        return r + 1

    r = 1
    inf.cell(r, 1, "CURRENT STATE - how it's built").font = Font(bold=True, size=13)
    r += 1
    inf.cell(r, 1, f"Scheduled jobs: {facts.get('scheduled_jobs')}  |  "
                   f"Admin commands: {facts.get('admin_commands')}  |  "
                   f"Vendors: {facts.get('vendors')}  |  Catalog items: "
                   f"{facts.get('catalog_items')}  |  Test files: {facts.get('test_files')}")
    r += 2
    r = _section(inf, r, "Architecture", CURRENT_STATE)
    r = _section(inf, r, "AI WORKLOAD ASSESSMENT (use / frequency / cost)", AI_WORKLOAD)
    r = _section(inf, r, "FUTURE-STATE RECOMMENDATIONS", FUTURE_STATE)
    r = _section(inf, r, "ADDITIONAL BUSINESS OPPORTUNITIES", BUSINESS_OPPS)
    r = _section(inf, r, "COST-REDUCTION OPPORTUNITIES", COST_REDUCTION)
    inf.column_dimensions["A"].width = 32
    inf.column_dimensions["B"].width = 90

    # ── Tab 5: Workflow (Etsy order -> delivery) ──
    try:
        from quoteforge.etsy.workflow_doc import STAGES
        wf = wb.create_sheet("Workflow")
        wf.append(["Stage", "Customer experience", "Behind the scenes", "AI",
                   "Status"])
        for cc in wf[1]:
            cc.font = BOLDW; cc.fill = HDR
        for st in STAGES:
            wf.append(list(st))
        for col, w in zip("ABCDE", (16, 34, 46, 30, 9)):
            wf.column_dimensions[col].width = w
        for row in wf.iter_rows(min_row=2):
            for cc in row:
                cc.alignment = Alignment(wrap_text=True, vertical="top")
    except Exception:  # noqa: BLE001
        pass

    from pathlib import Path
    out = Path(out_path) if out_path else (OUTPUT_DIR / "executive_report.xlsx")
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    return out


def _ai_exec_summary(shop: str, k: dict, facts: dict) -> str:
    """AI-written executive summary (Claude) with a deterministic fallback."""
    import json
    from quoteforge.ai.assistant import ai_text
    fallback = (
        f"{shop} is a fully-automated personalized wall-art business on Etsy + "
        f"Gelato. Month-to-date revenue ${k['revenue']} at {k['margin_pct']}% net "
        f"margin across {k['orders']} orders (AOV ${k['aov']}). The platform runs "
        f"{facts.get('scheduled_jobs')} scheduled jobs and {facts.get('admin_commands')} "
        f"commands end-to-end: AI quote generation, QC-before-proof, vendor-routed "
        f"fulfillment, tracking, retention and reporting - with a 60% margin floor "
        f"and human-approved refunds. Near-term: drive traffic (Pinterest/SEO/"
        f"email), grow reviews, and host the server for live AI support. "
        f"Biggest upside: subscriptions, B2B/wholesale, and digital downloads.")
    return ai_text(
        f"Write a 6-8 sentence executive summary for {shop} (personalized wall-art "
        f"POD on Etsy+Gelato) from these metrics + infra facts. Cover performance, "
        f"automation maturity, and the top 3 growth + 1 cost-saving move.\n\n"
        f"METRICS: {json.dumps(k, default=str)}\nINFRA: {json.dumps(facts, default=str)}",
        operation="exec_summary", max_tokens=420, mock=fallback)
