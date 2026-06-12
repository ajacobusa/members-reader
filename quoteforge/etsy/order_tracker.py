"""Order tracking Excel exporter.

Reads quoteforge/order_log.csv and exports a color-coded, formula-driven
Excel workbook matching the Custom_Quote_Order_Workflow.xlsx column layout:
  Order ID | Customer Name | Occasion | Recipient | Scenery |
  Quote Status | Design Status | Gelato Upload | Tracking Number | Notes
"""
import csv
import uuid
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule, ColorScaleRule, DataBarRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

from quoteforge.config import OUTPUT_DIR

# ── Color palette ──────────────────────────────────────────────
CLR = {
    "header_bg":   "1F4E79",   # dark blue header
    "header_fg":   "FFFFFF",
    "alt_row":     "EBF3FB",   # light blue alt rows
    "white":       "FFFFFF",
    "pending":     "FFF2CC",   # yellow — pending
    "in_progress": "DEEBF7",   # light blue — in progress
    "done":        "E2EFDA",   # green — done
    "border":      "BDD7EE",
    "dash_header": "2E75B6",   # dashboard header blue
    "dash_bg":     "D6E4F0",
}

# ── Status dropdown options ────────────────────────────────────
QUOTE_STATUSES  = ["Pending", "Generated", "Reviewed", "Approved"]
DESIGN_STATUSES = ["Not Started", "In Canva", "Exported", "Done"]
GELATO_STATUSES = ["Not Uploaded", "Uploaded", "In Production", "Shipped"]

COLUMNS = [
    ("Order ID",       14),
    ("Customer Name",  20),
    ("Occasion",       22),
    ("Recipient",      18),
    ("Scenery",        18),
    ("Quote Status",   16),
    ("Design Status",  16),
    ("Gelato Upload",  16),
    ("Tracking #",     18),
    ("Notes",          30),
]
NCOLS = len(COLUMNS)


def _thin_border(color: str = "CCCCCC") -> Border:
    """Thin border on all four sides in the given color."""
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)


def _cell_fill(hex_color: str) -> PatternFill:
    """Solid cell fill in the given hex color."""
    return PatternFill("solid", fgColor=hex_color)


def _hdr_font(bold: bool = True) -> Font:
    """Header font in the tracker's header color."""
    return Font(name="Arial", bold=bold, color=CLR["header_fg"], size=11)


def _body_font(bold: bool = False) -> Font:
    """Standard body font for data cells."""
    return Font(name="Arial", bold=bold, size=10)


def _read_order_log() -> list[dict]:
    """Read order_log.csv rows as dicts; empty list if the log doesn't exist."""
    log_path = OUTPUT_DIR / "order_log.csv"
    if not log_path.exists():
        return []
    with log_path.open(encoding="utf-8") as f:
        return list(csv.DictReader(f))


def export_order_tracker(output_path: Path | None = None) -> Path:
    """Export order log to a formatted Excel tracker workbook.

    Returns the path to the saved .xlsx file.
    """
    if output_path is None:
        output_path = OUTPUT_DIR / "QuoteForge_Order_Tracker.xlsx"

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    log_rows = _read_order_log()

    wb = Workbook()

    # ── Sheet 1: Orders ─────────────────────────────────────────
    ws = wb.active
    ws.title = "Orders"
    ws.freeze_panes = "A2"  # freeze header row

    # Header row
    for col_idx, (col_name, col_width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = _hdr_font()
        cell.fill = _cell_fill(CLR["header_bg"])
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _thin_border(CLR["border"])
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    ws.row_dimensions[1].height = 30

    # Data rows
    for row_num, log in enumerate(log_rows, start=2):
        # Generate stable Order ID from timestamp + recipient
        order_id = f"QF-{row_num-1:04d}"
        row_data = [
            order_id,
            log.get("sender_name", ""),          # customer = sender (buyer)
            log.get("occasion", ""),
            log.get("recipient_name", ""),
            log.get("scenery", ""),
            "Generated",                          # quote was already generated
            "Not Started",                        # design starts manually
            "Not Uploaded",                       # gelato starts manually
            "",                                   # tracking number — filled later
            log.get("output_style", ""),          # notes = style used
        ]
        alt = (row_num % 2 == 0)
        row_fill = _cell_fill(CLR["alt_row"] if alt else CLR["white"])

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_num, column=col_idx, value=value)
            cell.font = _body_font()
            cell.fill = row_fill
            cell.alignment = Alignment(vertical="center", wrap_text=False)
            cell.border = _thin_border(CLR["border"])

    # If no rows yet, add 30 blank placeholder rows
    start_blank = len(log_rows) + 2
    if not log_rows:
        start_blank = 2
    for row_num in range(start_blank, start_blank + 30):
        alt = (row_num % 2 == 0)
        row_fill = _cell_fill(CLR["alt_row"] if alt else CLR["white"])
        order_id = f"QF-{row_num-1:04d}"
        ws.cell(row=row_num, column=1, value=order_id).fill = row_fill
        for col_idx in range(2, NCOLS + 1):
            cell = ws.cell(row=row_num, column=col_idx, value="")
            cell.fill = row_fill
            cell.border = _thin_border(CLR["border"])
            cell.font = _body_font()

    last_data_row = max(start_blank + 29, len(log_rows) + 1)

    # ── Dropdowns (data validation) ──────────────────────────────
    def _add_dropdown(col: int, options: list[str]) -> None:
        """Attach a list data-validation dropdown to a column's data rows."""
        formula = f'"{",".join(options)}"'
        dv = DataValidation(type="list", formula1=formula, allow_blank=True, showDropDown=False)
        dv.sqref = f"{get_column_letter(col)}2:{get_column_letter(col)}{last_data_row}"
        ws.add_data_validation(dv)

    _add_dropdown(6, QUOTE_STATUSES)
    _add_dropdown(7, DESIGN_STATUSES)
    _add_dropdown(8, GELATO_STATUSES)

    # ── Conditional formatting: status colors ────────────────────
    def _status_rule(col: int, value: str, fill_hex: str) -> None:
        """Color cells in a column when they contain the given status text."""
        col_letter = get_column_letter(col)
        fill = PatternFill(start_color=fill_hex, end_color=fill_hex, fill_type="solid")
        font = Font(name="Arial", size=10, bold=True)
        dstyle = DifferentialStyle(fill=fill, font=font)
        rule = Rule(type="containsText", operator="containsText",
                    text=value, dxf=dstyle)
        rule.formula = [f'NOT(ISERROR(SEARCH("{value}",{col_letter}2)))']
        ws.conditional_formatting.add(
            f"{col_letter}2:{col_letter}{last_data_row}", rule)

    # Quote status colors
    _status_rule(6, "Pending",     CLR["pending"])
    _status_rule(6, "Generated",   CLR["in_progress"])
    _status_rule(6, "Approved",    CLR["done"])

    # Design status colors
    _status_rule(7, "Not Started", CLR["pending"])
    _status_rule(7, "In Canva",    CLR["in_progress"])
    _status_rule(7, "Done",        CLR["done"])

    # Gelato upload colors
    _status_rule(8, "Not Uploaded",   CLR["pending"])
    _status_rule(8, "Uploaded",       CLR["in_progress"])
    _status_rule(8, "Shipped",        CLR["done"])

    # ── Sheet 2: Dashboard ───────────────────────────────────────
    dash = wb.create_sheet("Dashboard")
    dash.sheet_view.showGridLines = False

    def _dash_header(row: int, text: str) -> None:
        """Write a merged, styled section header on the dashboard sheet."""
        cell = dash.cell(row=row, column=2, value=text)
        cell.font = Font(name="Arial", bold=True, size=13, color=CLR["header_fg"])
        cell.fill = _cell_fill(CLR["dash_header"])
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _thin_border()
        dash.merge_cells(f"B{row}:E{row}")
        dash.row_dimensions[row].height = 24

    def _dash_row(row: int, label: str, formula: str) -> None:
        """Write one dashboard metric row: label plus a formula value cell."""
        lbl = dash.cell(row=row, column=2, value=label)
        lbl.font = _body_font(bold=True)
        lbl.fill = _cell_fill(CLR["dash_bg"])
        lbl.alignment = Alignment(horizontal="left", vertical="center")
        lbl.border = _thin_border()

        val = dash.cell(row=row, column=5, value=formula)
        val.font = Font(name="Arial", bold=True, size=12)
        val.fill = _cell_fill(CLR["white"])
        val.alignment = Alignment(horizontal="center", vertical="center")
        val.border = _thin_border()

        for c in [3, 4]:
            cell = dash.cell(row=row, column=c, value="")
            cell.fill = _cell_fill(CLR["dash_bg"])
            cell.border = _thin_border()

        dash.row_dimensions[row].height = 22

    for c in ["B", "C", "D", "E"]:
        dash.column_dimensions[c].width = 22

    _dash_header(2, "QuoteForge Order Dashboard")

    _dash_header(4, "Order Summary")
    _dash_row(5, "Total Orders",
              f"=COUNTA(Orders!A2:A{last_data_row})-COUNTIF(Orders!A2:A{last_data_row},\"\")")
    _dash_row(6, "Quotes Pending",
              f"=COUNTIF(Orders!F2:F{last_data_row},\"Pending\")")
    _dash_row(7, "Quotes Generated",
              f"=COUNTIF(Orders!F2:F{last_data_row},\"Generated\")")
    _dash_row(8, "Quotes Approved",
              f"=COUNTIF(Orders!F2:F{last_data_row},\"Approved\")")

    _dash_header(10, "Design Status")
    _dash_row(11, "Not Started",
              f"=COUNTIF(Orders!G2:G{last_data_row},\"Not Started\")")
    _dash_row(12, "In Canva",
              f"=COUNTIF(Orders!G2:G{last_data_row},\"In Canva\")")
    _dash_row(13, "Design Done",
              f"=COUNTIF(Orders!G2:G{last_data_row},\"Done\")")

    _dash_header(15, "Gelato Fulfillment")
    _dash_row(16, "Not Uploaded",
              f"=COUNTIF(Orders!H2:H{last_data_row},\"Not Uploaded\")")
    _dash_row(17, "Uploaded / In Production",
              f"=COUNTIF(Orders!H2:H{last_data_row},\"Uploaded\")+COUNTIF(Orders!H2:H{last_data_row},\"In Production\")")
    _dash_row(18, "Shipped",
              f"=COUNTIF(Orders!H2:H{last_data_row},\"Shipped\")")

    _dash_header(20, "Workflow — 14 Steps")
    steps = [
        "1. Create Etsy seller account + banking/tax",
        "2. Create Gelato account + payment",
        "3. Connect Etsy → Gelato",
        "4. Create poster/canvas products in Gelato",
        "5. Design templates in Canva (50-100 masters)",
        "6. Create niche Etsy listings w/ personalization box",
        "7. Receive order → copy customer details",
        "8. Generate custom quote in QuoteForge (Order tab)",
        "9. Insert quote into Canva template",
        "10. Export PNG at 300 DPI",
        "11. Upload PNG to Gelato → approve",
        "12. Gelato prints + ships automatically",
        "13. Send follow-up email (QuoteForge Order tab)",
        "14. Scale: SEO, Pinterest, product expansion",
    ]
    for i, step in enumerate(steps, start=21):
        cell = dash.cell(row=i, column=2, value=step)
        cell.font = Font(name="Arial", size=10)
        cell.fill = _cell_fill(CLR["alt_row"] if i % 2 == 0 else CLR["white"])
        cell.border = _thin_border()
        dash.merge_cells(f"B{i}:E{i}")
        dash.row_dimensions[i].height = 18

    # ── Sheet 3: SEO Reference ───────────────────────────────────
    seo = wb.create_sheet("SEO Reference")
    seo.sheet_view.showGridLines = False
    seo_data = [
        ("Niche", "Example Title (≤140 chars)", "Key Tags"),
        ("Daughter Gift", "Personalized Daughter Gift | Custom Quote Print | Graduation Gift For Daughter", "daughter gift, graduation gift, custom wall art, mountain decor, quote poster"),
        ("Son Gift", "Personalized Son Gift | Custom Motivational Print | Gift For Son", "son gift, graduation gift, dad to son, personalized art, office decor"),
        ("Wife Gift", "To My Wife | Custom Love Letter Print | Anniversary Gift", "wife gift, anniversary gift, love letter print, romantic gift, valentines day"),
        ("Husband Gift", "To My Husband | Custom Love Letter Print | Anniversary Gift", "husband gift, anniversary gift, love letter print, romantic gift"),
        ("Mom Gift", "Personalized Mom Gift | Custom Quote From Daughter | Mother's Day", "mom gift, mothers day, custom wall art, gift for mom, flower decor"),
        ("Dad Gift", "Personalized Dad Gift | Custom Quote From Daughter | Father's Day", "dad gift, fathers day, custom wall art, gift for dad, mountain decor"),
        ("Christian Art", "Christian Wall Art | Faith Encouragement Print | Bible Inspired", "christian gift, faith wall art, christian decor, bible inspired, prayer art"),
        ("Graduation", "Graduation Gift | Custom Quote Print | Class of 2026 Wall Art", "graduation gift, class of 2026, motivational print, college grad"),
        ("Memorial", "Memorial Gift | Personalized Remembrance Print | In Memory Wall Art", "memorial gift, sympathy gift, in memory of, grief healing"),
        ("Future Dentist", "Future Dentist Gift | Dental School Motivation | DAT Study Decor", "future dentist, dental school, dat motivation, pre-dental gift"),
        ("Pet Memorial", "Pet Memorial Print | Dog Rainbow Bridge | Custom Pet Remembrance", "pet memorial, rainbow bridge, dog memorial, cat memorial, pet gift"),
    ]
    seo_widths = [20, 70, 60]
    for col_idx, width in enumerate(seo_widths, start=1):
        seo.column_dimensions[get_column_letter(col_idx)].width = width

    for row_idx, row_data in enumerate(seo_data, start=1):
        for col_idx, value in enumerate(row_data, start=1):
            cell = seo.cell(row=row_idx, column=col_idx, value=value)
            cell.border = _thin_border()
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if row_idx == 1:
                cell.font = Font(name="Arial", bold=True, size=11,
                                 color=CLR["header_fg"])
                cell.fill = _cell_fill(CLR["header_bg"])
            else:
                cell.font = _body_font()
                cell.fill = _cell_fill(
                    CLR["alt_row"] if row_idx % 2 == 0 else CLR["white"])
        seo.row_dimensions[row_idx].height = 30 if row_idx == 1 else 40

    wb.save(output_path)
    return output_path
