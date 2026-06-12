"""Monthly reconciliation Excel export.

Produces a bookkeeping ledger: every order with sale price, Etsy fees, sales
tax (pass-through), Gelato cost, and net profit — plus totals. Use it to
reconcile against your Etsy statement, Gelato charges, and bank deposits.
"""
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from quoteforge.config import OUTPUT_DIR
from quoteforge.etsy.financials import month_financials

HDR_BG = "1F4E79"
TOTAL_BG = "D6E4F0"


def _border():
    """Thin light-grey border on all four sides."""
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def export_reconciliation(year: int, month: int,
                          output_path: Path | None = None) -> Path:
    """Export a monthly reconciliation ledger to Excel. Returns the path."""
    data = month_financials(year, month)
    if output_path is None:
        OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
        output_path = OUTPUT_DIR / f"reconciliation_{data['period']}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Reconciliation"

    # Title
    ws.merge_cells("A1:H1")
    t = ws["A1"]
    t.value = f"QuoteForge Financial Reconciliation — {data['period']}"
    t.font = Font(name="Arial", bold=True, size=14, color="FFFFFF")
    t.fill = PatternFill("solid", fgColor=HDR_BG)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    headers = ["Order ID", "Etsy Order", "Date", "Occasion", "Sale Price ($)",
               "Etsy Fees ($)", "Sales Tax ($, remitted by Etsy)",
               "Gelato Cost ($)", "Net Profit ($)"]
    widths = [16, 16, 12, 20, 14, 13, 24, 14, 14]
    # Note: 9 columns; widen merge to A1:I1
    ws.unmerge_cells("A1:H1")
    ws.merge_cells("A1:I1")

    for col, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(row=2, column=col, value=h)
        c.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill("solid", fgColor=HDR_BG)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _border()
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[2].height = 32

    row = 3
    for r in data["rows"]:
        vals = [
            r["order_id"], r["etsy_order_id"], r["created_at"], r["occasion"],
            r["sale_price"], r["etsy_fees"], r["sales_tax_collected"],
            r["gelato_cost"], r["net_profit"],
        ]
        alt = (row % 2 == 0)
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=row, column=col, value=v)
            c.font = Font(name="Arial", size=10)
            c.fill = PatternFill("solid", fgColor="F2F2F2" if alt else "FFFFFF")
            c.border = _border()
            if col >= 5:
                c.number_format = "#,##0.00"
        row += 1

    # Totals row
    totals = ["TOTAL", "", "", f"{data['order_count']} orders",
              data["revenue"], data["etsy_fees"], data["sales_tax_collected"],
              data["gelato_cost"], data["net_profit"]]
    for col, v in enumerate(totals, 1):
        c = ws.cell(row=row, column=col, value=v)
        c.font = Font(name="Arial", bold=True, size=11)
        c.fill = PatternFill("solid", fgColor=TOTAL_BG)
        c.border = _border()
        if col >= 5:
            c.number_format = "#,##0.00"

    # Reconciliation notes
    note_row = row + 2
    notes = [
        "How to reconcile:",
        f"1. Etsy deposits (money IN) should match: Revenue − Etsy Fees = "
        f"${round(data['revenue'] - data['etsy_fees'], 2)}",
        f"2. Gelato charges (money OUT) should match: ${data['gelato_cost']}",
        f"3. Your true profit for the month: ${data['net_profit']}",
        "Sales tax is collected AND remitted by Etsy — it is not your income or a cost.",
        "Pending/error orders are excluded (no revenue earned yet).",
    ]
    for i, n in enumerate(notes):
        c = ws.cell(row=note_row + i, column=1, value=n)
        c.font = Font(name="Arial", size=10, bold=(i == 0))
        ws.merge_cells(start_row=note_row + i, start_column=1,
                       end_row=note_row + i, end_column=9)

    wb.save(output_path)
    return output_path
