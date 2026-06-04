"""Seasonal campaign generator — batch-create a month's listings, each with a
recommended PUBLISH-BY date so you rank in Etsy search before buyers shop.

Why timing matters: a new Etsy listing takes ~3-6 weeks to mature in search
ranking. To be first in front of buyers for a holiday, you must publish weeks
before the peak. This module computes that publish-by date per occasion and
produces ready-to-build listing plans (title, tags, scenery, quote hint),
sorted most-urgent first.
"""
from datetime import datetime, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from quoteforge.config import OUTPUT_DIR
from quoteforge.etsy.occasions import get_month_occasions

# Default lead time: publish ~5 weeks before the peak so search ranking matures.
DEFAULT_LEAD_DAYS = 35

# Fixed-date holiday peaks (month, day). Variable ones use mid-month as a proxy.
HOLIDAY_PEAKS: dict[str, tuple[int, int]] = {
    "Valentine's Day": (2, 14),
    "Galentine's Day": (2, 13),
    "Independence Day": (7, 4),
    "Halloween": (10, 31),
    "Veterans Day": (11, 11),
    "Christmas": (12, 25),
    "New Year's Day": (1, 1),
}

# Light scenery + quote-category hints by occasion keyword (for the designer/AI).
SCENERY_HINTS: list[tuple[str, str]] = [
    ("memorial", "Soft Sunrise / Gentle"),
    ("loss", "Soft Sunrise / Gentle"),
    ("grief", "Soft Sunrise / Gentle"),
    ("wedding", "Floral / Warm Bokeh"),
    ("anniversary", "Floral / Warm Bokeh"),
    ("bride", "Floral / Warm Bokeh"),
    ("groom", "Floral / Warm Bokeh"),
    ("graduation", "Mountain Sunrise"),
    ("future", "Mountain Sunrise"),
    ("baptism", "Golden Light"),
    ("confirmation", "Golden Light"),
    ("communion", "Golden Light"),
    ("christian", "Golden Light"),
    ("faith", "Golden Light"),
    ("christmas", "Cozy Winter"),
    ("baby", "Soft Pastel"),
    ("pregnancy", "Soft Pastel"),
    ("retirement", "Beach / Open Road"),
    ("military", "Flag / Patriotic"),
    ("veteran", "Flag / Patriotic"),
    ("independence", "Flag / Patriotic"),
]


def _scenery_for(occasion: str) -> str:
    low = occasion.lower()
    for key, scenery in SCENERY_HINTS:
        if key in low:
            return scenery
    return "Mountains"


def recommended_publish_date(occasion: str, target_month: int,
                             now: datetime | None = None,
                             lead_days: int = DEFAULT_LEAD_DAYS) -> dict:
    """Compute when to PUBLISH this listing to rank before its peak.

    Returns the peak date, the publish-by date, days remaining, and an urgency
    flag (overdue / urgent / on-track).
    """
    now = now or datetime.now()
    year = now.year
    if target_month < now.month:
        year += 1  # the peak is next year (e.g. planning Jan in December)

    if occasion in HOLIDAY_PEAKS:
        pm, pd = HOLIDAY_PEAKS[occasion]
        peak = datetime(year, pm, pd)
    else:
        peak = datetime(year, target_month, 15)  # mid-month proxy peak

    publish_by = peak - timedelta(days=lead_days)
    days_to_publish = (publish_by.date() - now.date()).days
    if days_to_publish < 0:
        urgency = "OVERDUE — list ASAP"
    elif days_to_publish <= 7:
        urgency = "URGENT — list this week"
    else:
        urgency = "On track"
    return {
        "peak": peak.strftime("%Y-%m-%d"),
        "publish_by": publish_by.strftime("%Y-%m-%d"),
        "days_to_publish": days_to_publish,
        "urgency": urgency,
    }


def build_listing_plan(occasion: str, target_month: int,
                       now: datetime | None = None) -> dict:
    """A ready-to-build listing plan for one occasion."""
    scenery = _scenery_for(occasion)
    timing = recommended_publish_date(occasion, target_month, now)

    # SEO title (≤140 chars) and tags from the occasion
    title = (f"Personalized {occasion} Gift | Custom Quote Wall Art | "
             f"Scenic {scenery.split(' /')[0]} Poster Print")[:140]
    base_words = [w for w in occasion.lower().replace("—", " ").split()
                  if len(w) > 2][:4]
    tags = []
    for w in base_words:
        tags.append(f"{w} gift"[:20])
    tags += ["custom wall art", "personalized gift", "quote poster",
             "scenic print", "meaningful gift", "custom quote", "wall decor",
             "gift idea", "keepsake print"]
    tags = list(dict.fromkeys(t for t in tags if len(t) <= 20))[:13]

    return {
        "occasion": occasion,
        "scenery": scenery,
        "title": title,
        "tags": tags,
        "publish_by": timing["publish_by"],
        "peak": timing["peak"],
        "days_to_publish": timing["days_to_publish"],
        "urgency": timing["urgency"],
        "quote_hint": f"A heartfelt, original message for: {occasion}",
    }


def seasonal_campaign(month: int | str, now: datetime | None = None) -> list[dict]:
    """All listing plans for a month's occasions, MOST URGENT first."""
    now = now or datetime.now()
    if isinstance(month, str):
        month_num = datetime.strptime(month, "%B").month
    else:
        month_num = month
    occasions = get_month_occasions(month_num)
    plans = [build_listing_plan(o, month_num, now) for o in occasions]
    plans.sort(key=lambda p: p["days_to_publish"])  # soonest publish-by first
    return plans


def export_campaign_excel(month: int | str, output_path: Path | None = None,
                          now: datetime | None = None) -> Path:
    """Export a month's campaign plan (with publish-by dates) to Excel."""
    now = now or datetime.now()
    if isinstance(month, str):
        month_name = month
    else:
        month_name = datetime(2000, month, 1).strftime("%B")
    plans = seasonal_campaign(month, now)

    if output_path is None:
        OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
        output_path = OUTPUT_DIR / f"campaign_{month_name}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Campaign"
    border = Border(*([Side(style="thin", color="CCCCCC")] * 4))

    ws.merge_cells("A1:G1")
    t = ws["A1"]
    t.value = f"QuoteForge {month_name} Campaign — publish early to rank first"
    t.font = Font(name="Arial", bold=True, size=14, color="FFFFFF")
    t.fill = PatternFill("solid", fgColor="1F4E79")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    headers = ["Publish By", "Urgency", "Occasion", "Scenery",
               "Etsy Title", "Tags", "Peak Date"]
    widths = [13, 20, 24, 18, 60, 50, 12]
    for col, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(row=2, column=col, value=h)
        c.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill("solid", fgColor="1F4E79")
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        c.border = border
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[2].height = 28

    for i, p in enumerate(plans, start=3):
        urgent = "OVERDUE" in p["urgency"] or "URGENT" in p["urgency"]
        bg = "FCE4D6" if urgent else ("F2F2F2" if i % 2 == 0 else "FFFFFF")
        vals = [p["publish_by"], p["urgency"], p["occasion"], p["scenery"],
                p["title"], ", ".join(p["tags"]), p["peak"]]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=i, column=col, value=v)
            c.font = Font(name="Arial", size=9,
                          bold=(col == 2 and urgent))
            c.fill = PatternFill("solid", fgColor=bg)
            c.alignment = Alignment(vertical="center", wrap_text=(col in (5, 6)))
            c.border = border

    wb.save(output_path)
    return output_path
